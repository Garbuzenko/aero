# parser_file.py
"""
Модуль импорта и парсинга данных.
Загружает файлы с FTP, парсит данные и заполняет таблицу processed_flights.
"""
import mysql.connector
from mysql.connector import Error
import re
import logging
from datetime import datetime, date, time
import time as time_module
from typing import Optional, Dict, Any, List
import json
import ast
from shapely.geometry import Point, Polygon
from shapely.strtree import STRtree
from shapely.prepared import prep
import pandas as pd
from ftplib import FTP
import os
# from utils import settings
from openpyxl import load_workbook
import sys
import os
import re
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import settings

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('../log.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)


class FlightDataProcessor:
    # Список столбцов для формирования уникальной строки UNIQ_STR
    # Легко изменяемый список для анализа дубликатов
    UNIQ_STR_COLUMNS = [
        'dof_date',
        'registration',
        'operator',
        'aircraft_type_code',
        'filename',
        'region_id',
        'prediction',
        'sts',
        'time_of_day',
        'departure_coords'
    ]
    
    def __init__(self):
        self.connection = None
        self.regions = []
        self.region_tree = None
        self.sheet_structures = {}
        self.batch_data = []
        self.local_path = ""
        self.current_filename = None
        self.current_file_progress = 0
        self.total_records_in_file = 0

    def clean_line_breaks(self, text: str) -> str:
        """
        Удаление переносов строк из текста (замена на пробелы).
        Используется для очистки полей SHR, DEP, ARR перед обработкой.
        """
        if not text or not isinstance(text, str):
            return ""
        
        # Удаляем все виды переносов строк и заменяем их на пробелы
        text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        # Убираем множественные пробелы
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def extract_operator(self, text: str) -> str:
        """
        Извлечение оператора из текста с улучшенной обработкой многострочных блоков.
        Аналог PHP функции extractOperator.
        Улучшенная версия на основе кода друга.
        """
        if not text or not isinstance(text, str):
            return ""

        # Удаление переносов строк - заменяем на пробелы
        text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        
        # Разбиваем текст на строки (после замены переносов это будет одна строка,
        # но сохраняем логику для совместимости)
        lines = text.split('\n')
        
        in_operator_block = False
        operator_content = ''

        for line in lines:
            line = line.strip()

            # Если находим начало OPR/ блока
            opr_match = re.search(r'OPR/(.*)$', line, re.IGNORECASE)
            if opr_match:
                in_operator_block = True
                operator_content = opr_match.group(1)
                continue

            # Если мы внутри блока оператора
            if in_operator_block:
                # Проверяем, не начался ли следующий блок
                if (re.match(r'^[A-Z]{2,}/', line) or
                        re.match(r'^[A-Z]{3,}\s+[A-Z]+/', line)):
                    in_operator_block = False
                else:
                    # Добавляем содержимое строки к оператору
                    operator_content += ' ' + line

        # Очищаем от возможных хвостов следующих блоков
        if operator_content:
            # Удаляем хвосты следующих блоков
            operator_content = re.sub(r'\s+[A-Z]{2,}/.*$', '', operator_content, flags=re.IGNORECASE)
            operator_content = re.sub(r'\s+[A-Z]{3,}\s+[A-Z]+/.*$', '', operator_content, flags=re.IGNORECASE)

            # Удаляем телефонные номера (российский формат)
            # Паттерн совпадает с PHP версией
            operator_content = re.sub(
                r'(?:\+7|8)?[\s\-\(\)]*\d{3}[\s\-\(\)]*\d{3}[\s\-\(\)]*\d{2}[\s\-\(\)]*\d{2}',
                '',
                operator_content
            )
            
            # Убираем множественные пробелы
            operator_content = re.sub(r'\s+', ' ', operator_content)

            return operator_content.strip()

        return ""

    def extract_sts(self, text: str) -> Optional[str]:
        """
        Извлечение причины особого отношения (STS) из текста сообщения.
        
        Возвращаемые значения:
        - FFR - борьба с пожаром
        - SAR - гражданская оборона, чрезвычайные ситуации, спасательные операции
        - STATE - оборона и безопасность РФ
        - Другие значения из RMK/
        """
        if not text or not isinstance(text, str):
            return None
        
        # Ищем STS/ с последующим значением
        sts_match = re.search(r'STS/(\w+)', text, re.IGNORECASE)
        if sts_match:
            sts_value = sts_match.group(1).upper()
            # Проверяем на известные значения
            if sts_value in ['FFR', 'SAR', 'STATE']:
                return sts_value
            else:
                return sts_value
        
        # Ищем в блоке RMK/ после STS
        rmk_sts_match = re.search(r'RMK/.*?STS[/\s]+(\w+)', text, re.IGNORECASE | re.DOTALL)
        if rmk_sts_match:
            sts_value = rmk_sts_match.group(1).upper()
            return sts_value
        
        return None

    # НОВЫЕ УЛУЧШЕННЫЕ ФУНКЦИИ ДЛЯ РЕГИСТРАЦИОННЫХ НОМЕРОВ
    def extract_registration_numbers(self, text: str) -> List[str]:
        """Основная функция извлечения регистрационных номеров (аналог PHP)"""
        results = []

        # Сначала ищем в блоке REG/ - это приоритетный источник
        reg_results = self.extract_from_reg_block(text)

        # Если в REG/ нашли номера, возвращаем их (доверяем REG/ блоку)
        if reg_results:
            return reg_results

        # Если в REG/ ничего не нашли, смотрим TYP/ для определения количества БВС
        bvs_count = self.get_bvs_count_from_typ(text)

        if bvs_count > 0:
            # Ищем номера в других местах
            additional_results = self.find_registration_numbers_in_text(text, bvs_count)
            results.extend(additional_results)

        # Удаляем дубликаты и переиндексируем
        results = list(dict.fromkeys(results))

        return results

    def extract_from_reg_block(self, text: str) -> List[str]:
        """Извлекает номера из блока REG/ с улучшенной обработкой переносов строк"""
        results = []

        # Улучшенный паттерн для REG/ блока - ищем всё до следующего блока
        pattern = r'REG/([^\r\n]*(?:\s+[^\r\n\/]+)*)'

        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            reg_block = match.group(1).strip()

            # Удаляем возможные хвосты следующих блоков
            reg_block = re.sub(r'\s+[A-Z]{2,}/.*$', '', reg_block, flags=re.IGNORECASE | re.DOTALL)
            reg_block = re.sub(r'\s+[A-Z]{3,} .*$', '', reg_block, flags=re.IGNORECASE | re.DOTALL)

            # Обрабатываем блок - разбиваем по всем возможным разделителям
            parts = re.split(r'[\s,]+', reg_block)

            for part in parts:
                part = part.strip()
                if part:
                    # Очищаем от префикса REG если есть
                    clean_number = re.sub(r'^REG', '', part, flags=re.IGNORECASE)

                    # Минимальная проверка для номеров из REG/ блока
                    if self.is_valid_reg_number(clean_number):
                        results.append(clean_number)

        return results

    def is_valid_reg_number(self, number: str) -> bool:
        """Минимальная проверка для номеров из REG/ блока"""
        if not number or len(number) < 4 or len(number) > 12:
            return False

        # Должен состоять только из букв, цифр и дефиса
        if not re.match(r'^[A-Z0-9\-]+$', number, re.IGNORECASE):
            return False

        # Базовые исключения
        invalid_patterns = [
            r'^\d{6}$',  # Даты (250102)
            r'^[78]\d{9,10}$',  # Телефоны
            r'^SID\/\d+$',  # SID номера
            r'^M\d{4}',  # M0000
            r'^[A-Z]{4}\d+$',  # ZZZZ0700
            r'^EET\/[A-Z0-9]+$',  # EET/блоки
            r'^WR\d+$',  # WR16594
            r'^FL\d+$',  # FL060, FL200
            r'^\d+-\d+$',  # 1500-FL200
            r'^\d+-\w+\d+$',  # 1500-FL200
            r'^DEP\/[A-Z0-9]+$',  # DEP/блоки
            r'^DEST\/[A-Z0-9]+$',  # DEST/блоки
        ]

        for pattern in invalid_patterns:
            if re.match(pattern, number, re.IGNORECASE):
                return False

        # Дополнительная проверка: не должен быть координатой
        if self.is_coordinate(number):
            return False

        return True

    def is_coordinate(self, number: str) -> bool:
        """Проверяет, является ли строка координатой"""
        # Паттерны для координат
        coordinate_patterns = [
            r'^\d+[NS]\d+[EW]$',  # 6004N03037E
            r'^\d+[NS]\d+[EW]\d+$',  # 600400N0303700E
            r'^[NS]\d+[EW]$',  # N07358E
        ]

        for pattern in coordinate_patterns:
            if re.match(pattern, number, re.IGNORECASE):
                return True

        return False

    def is_valid_registration_number(self, number: str) -> bool:
        """Улучшенная проверка валидности регистрационного номера для поиска в тексте"""
        if not number or len(number) < 4 or len(number) > 12:
            return False

        # Должен состоять только из букв, цифр и дефиса
        if not re.match(r'^[A-Z0-9\-]+$', number, re.IGNORECASE):
            return False

        # Список явно невалидных паттернов
        invalid_patterns = [
            r'^\d{6}$',  # Даты (250102)
            r'^\d{1,2}:\d{2}$',  # Время
            r'^[78]\d{9,10}$',  # Телефоны
            r'^SID\/\d+$',  # SID номера
            r'^WR\d+$',  # WR номера
            r'^\d{1,4}[ММ]?$',  # Высоты (300М)
            r'^[A-Z]{4}\d+$',  # ZZZZ0700
            r'^M\d{4}',  # M0000
            r'^K\d{4}',  # K0300
            r'^S\d{4}',  # S0610
            r'^\d+[-_]\d+$',  # 10-37
            r'^MR\d+$',  # MR0277
            r'^R\d+',  # R001
            r'^EET\/[A-Z0-9]+$',  # EET/блоки
            r'^WR\d+$',  # WR16594
            r'^FL\d+$',  # FL060, FL200
            r'^\d+-\w+\d+$',  # 1500-FL200
            r'^DEP\/[A-Z0-9]+$',  # DEP/блоки
            r'^DEST\/[A-Z0-9]+$',  # DEST/блоки
            r'^(DOF|EET|OPR|REG|TYP|RMK|STS|SID|DEP|DEST|ULLL|UUWV|USTV|USHH|USSV|USTR|USRR|USNN)',  # Ключевые слова
        ]

        for pattern in invalid_patterns:
            if re.match(pattern, number, re.IGNORECASE):
                return False

        # Должен содержать хотя бы одну цифру
        if not re.search(r'\d', number):
            return False

        # Для чисто цифровых номеров разрешаем только 4-6 цифр (11869, 00724 и т.д.)
        if re.match(r'^\d+$', number):
            return 4 <= len(number) <= 6

        # Дополнительная проверка: не должен быть координатой
        if self.is_coordinate(number):
            return False

        return True

    def find_registration_numbers_in_text(self, text: str, expected_count: int = 0, exclude: List[str] = None) -> List[
        str]:
        """Ищет номера во всем тексте"""
        if exclude is None:
            exclude = []

        results = []

        # Улучшенные паттерны для регистрационных номеров БВС
        patterns = [
            # Форматы типа RA-0938G, RF-37204
            r'\b([A-Z]{2}-\d{4,5}[A-Z]?)\b',

            # Форматы с буквами внутри: 079N076, 09K2404, 0S92698, B094841 и т.д.
            r'\b([A-Z0-9]{6,8})\b',

            # Цифровые форматы (4-6 цифр)
            r'\b(\d{4,6})\b',

            # Форматы типа RF37204
            r'\b([A-Z]{2}\d{5})\b'
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for number in matches:
                number = number.strip()
                if (self.is_valid_registration_number(number) and
                        number not in results and
                        number not in exclude):
                    results.append(number)

                    if expected_count > 0 and len(results) >= expected_count:
                        return results

        return results

    def get_bvs_count_from_typ(self, text: str) -> int:
        """Извлекает количество БВС из блока TYP/"""
        pattern = r'TYP/(\d*)[A-Z]+'

        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            number = match.group(1)
            return int(number) if number else 1

        return 0

    def extract_all_registrations(self, text: str) -> Dict[str, Any]:
        """Улучшенное извлечение всех регистрационных номеров БПЛА по аналогии с PHP коду"""
        if not text or not isinstance(text, str):
            return {'registration': None, 'registration_all': None}

        # Используем новый улучшенный метод
        registration_numbers = self.extract_registration_numbers(text)

        if registration_numbers:
            registration_all = ';'.join(registration_numbers)
            return {
                'registration': registration_numbers[0],  # Первый номер
                'registration_all': registration_all  # Все номера через точку с запятой
            }

        return {'registration': None, 'registration_all': None}

    # УЛУЧШЕННАЯ ФУНКЦИЯ ДЛЯ EET ИНФОРМАЦИИ
    def extract_eet_info(self, text: str) -> Optional[str]:
        """Извлечение расчетного времени полета (EET) с преобразованием в числовой формат (минуты)"""
        if not text or not isinstance(text, str):
            return None

        # Паттерны для поиска времени полета
        patterns = [
            # EET/0130 (часы+минуты)
            r'EET/(\d{4})\b',
            # EET 0130 (часы+минуты)
            r'EET\s+(\d{4})\b',
            # Время в формате ЧЧ:ММ
            r'EET/(\d{1,2}:\d{2})\b',
            # Число минут (30MIN, 120MIN)
            r'EET/(\d+)\s*MIN\b',
            r'EET/(\d+)\s*МИН\b',
            # Число часов (2H, 1.5H)
            r'EET/(\d+(?:\.\d+)?)\s*[HЧ]\b',
            # Время в тексте RMK
            r'RMK/.*?ВРЕМЯ[^\d]*(\d+(?:[\.:]\d+)?)\s*(?:МИН|MIN|Ч|H)',
            r'RMK/.*?ПОЛЕТА[^\d]*(\d+(?:[\.:]\d+)?)\s*(?:МИН|MIN|Ч|H)',
            # Общие паттерны для времени
            r'(\d{1,2}:\d{2})\s*(?:ЧАС|H|Ч)',
            r'(\d+(?:\.\d+)?)\s*(?:ЧАС|H|Ч)\s*(\d+)?\s*(?:МИН|MIN)?',
            r'(\d+)\s*(?:МИН|MIN)\b',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                for match in matches:
                    if isinstance(match, tuple):
                        # Обработка случаев типа (часы, минуты)
                        hours_str = match[0]
                        minutes_str = match[1] if len(match) > 1 else "0"
                    else:
                        hours_str = match
                        minutes_str = "0"

                    try:
                        # Обработка формата ЧЧММ (0130 = 1 час 30 минут)
                        if len(hours_str) == 4 and hours_str.isdigit():
                            hours = int(hours_str[:2])
                            minutes = int(hours_str[2:4])
                            if hours < 24 and minutes < 60:  # Проверяем валидность времени
                                total_minutes = hours * 60 + minutes
                                return str(total_minutes)

                        # Обработка формата ЧЧ:ММ
                        elif ':' in hours_str:
                            parts = hours_str.split(':')
                            if len(parts) == 2:
                                hours = int(parts[0])
                                minutes = int(parts[1])
                                total_minutes = hours * 60 + minutes
                                return str(total_minutes)

                        # Обработка десятичных часов (1.5 = 1 час 30 минут)
                        elif '.' in hours_str:
                            hours_float = float(hours_str)
                            hours = int(hours_float)
                            minutes = int((hours_float - hours) * 60)
                            total_minutes = hours * 60 + minutes
                            return str(total_minutes)

                        # Обработка только минут
                        elif hours_str.isdigit():
                            minutes_val = int(hours_str)
                            # Если число большое, вероятно это уже минуты
                            if minutes_val > 24:  # Более суток в часах маловероятно
                                return str(minutes_val)
                            else:
                                # Малое число может быть часами
                                if minutes_str.isdigit():
                                    additional_minutes = int(minutes_str)
                                else:
                                    additional_minutes = 0
                                total_minutes = minutes_val * 60 + additional_minutes
                                return str(total_minutes)

                    except (ValueError, TypeError) as e:
                        logging.debug(f"Ошибка парсинга времени '{match}': {e}")
                        continue

        # Дополнительный поиск в тексте примечаний
        if 'RMK/' in text:
            rmk_match = re.search(r'RMK/(.*?)(?:\s+[A-Z]{2,}/|$)', text, re.IGNORECASE | re.DOTALL)
            if rmk_match:
                rmk_text = rmk_match.group(1)
                # Поиск паттернов времени в примечаниях
                time_patterns = [
                    r'(\d+)\s*МИНУТ',
                    r'(\d+)\s*МИН',
                    r'(\d+)\s*MIN',
                    r'(\d+)\s*ЧАС',
                    r'(\d+)\s*H',
                    r'(\d+)\s*Ч',
                    r'(\d+)\s*\.\s*(\d+)\s*ЧАС',
                ]

                for pattern in time_patterns:
                    time_match = re.search(pattern, rmk_text, re.IGNORECASE)
                    if time_match:
                        try:
                            if len(time_match.groups()) == 2:
                                # Формат типа "1.5 ЧАС"
                                hours = int(time_match.group(1))
                                fraction = int(time_match.group(2))
                                total_minutes = hours * 60 + (fraction * 6)  # 0.5 часа = 30 минут
                            else:
                                # Формат с одним числом
                                num = int(time_match.group(1))
                                if any(word in time_match.group(0).upper() for word in ['ЧАС', 'H', 'Ч']):
                                    total_minutes = num * 60
                                else:
                                    total_minutes = num
                            return str(total_minutes)
                        except (ValueError, TypeError):
                            continue

        return None

    # СУЩЕСТВУЮЩИЕ ФУНКЦИИ (БЕЗ ИЗМЕНЕНИЙ)
    def process_single_flight(self, row: Dict[str, Any]):
        """Обработка одиночного полета с использованием новых методов парсинга."""
        try:
            # Извлекаем сообщения
            messages = self.extract_messages(row)
            if not any(messages.values()):
                return None

            # Парсим данные из всех типов сообщений
            shr_data = self.parse_shr_message(messages.get('shr', ''))
            dep_data = self.parse_dep_message(messages.get('dep', ''))
            arr_data = self.parse_arr_message(messages.get('arr', ''))

            # SID field removed from database schema

            # Сливаем данные
            flight_data = self.merge_flight_data(shr_data, dep_data, arr_data)

            # Дополнительное извлечение данных из исходных сообщений
            # если они не были найдены в парсированных данных
            if not flight_data.get('flight_number'):
                for msg_type, msg_text in messages.items():
                    if msg_text:
                        flight_number = self.extract_flight_number(msg_text)
                        if flight_number:
                            flight_data['flight_number'] = flight_number
                            break

            # ИСПОЛЬЗУЕМ УЛУЧШЕННУЮ ФУНКЦИЮ ДЛЯ EET
            if not flight_data.get('eet_info'):
                for msg_type, msg_text in messages.items():
                    if msg_text:
                        eet_info = self.extract_eet_info(msg_text)  # Используем улучшенную функцию
                        if eet_info:
                            flight_data['eet_info'] = eet_info
                            break

            if not flight_data.get('altitude_info'):
                for msg_type, msg_text in messages.items():
                    if msg_text:
                        altitude_info = self.extract_altitude_info(msg_text)
                        if altitude_info:
                            flight_data['altitude_info'] = altitude_info
                            break

            # Определение регионов
            region_id = 0
            departure_region = None
            arrival_region = None

            # Исправлено: было departure_lat дважды
            if flight_data.get('arrival_lat') and flight_data.get('arrival_lon'):
                result = self.find_region(flight_data['arrival_lat'], flight_data['arrival_lon'])
                if result:
                    region_id, arrival_region = result

            if flight_data.get('departure_lat') and flight_data.get('departure_lon'):  # Исправлено
                result = self.find_region(flight_data['departure_lat'], flight_data['departure_lon'])
                if result:
                    region_id, departure_region = result

            # Fallback логика для регионов
            if not arrival_region and departure_region:
                arrival_region = departure_region
            if not departure_region and arrival_region:
                departure_region = arrival_region


            # Fallback логика для координат
            if not flight_data.get('departure_lat') and flight_data.get('arrival_lat'):
                flight_data['departure_lat'] = flight_data['arrival_lat']
                flight_data['departure_lon'] = flight_data['arrival_lon']
                flight_data['departure_coords'] = flight_data['arrival_coords']
            if not flight_data.get('arrival_lat') and flight_data.get('departure_lat'):
                flight_data['arrival_lat'] = flight_data['departure_lat']
                flight_data['arrival_lon'] = flight_data['departure_lon']
                flight_data['arrival_coords'] = flight_data['departure_coords']


                     # Если region_id не был найден, но есть регионы, используем их
            if region_id == 0 and (departure_region or arrival_region):
                # Пытаемся найти region_id по названию региона
                for region in self.regions:
                    if region['name'] == departure_region or region['name'] == arrival_region:
                        region_id = region['id']
                        break

            # Сохраняем регионы в flight_data
            flight_data['region_id'] = region_id
            flight_data['departure_region'] = departure_region
            flight_data['arrival_region'] = arrival_region

            values = (
                str(row.get('source_id'))[:255] if row.get('source_id') else None,  # 1
                row.get('sheet_name'),  # 2
                flight_data.get('flight_id'),  # flight_id                 # 3
                flight_data.get('flight_number'),  # 4
                flight_data.get('aircraft_type_code'),  # 5
                flight_data.get('aircraft_type_desc'),  # 6
                flight_data.get('region_id', 0),  # 7
                flight_data.get('departure_region'),  # 8
                flight_data.get('arrival_region'),  # 9
                flight_data.get('duration_min'),  # 10
                flight_data.get('operator'),  # 11
                flight_data.get('registration'),  # 12
                flight_data.get('registration_all'),  # 13 ← УЛУЧШЕННОЕ поле!
                flight_data.get('departure_coords'),  # 14
                flight_data.get('arrival_coords'),  # 15
                flight_data.get('departure_lat'),  # 16
                flight_data.get('departure_lon'),  # 17
                flight_data.get('arrival_lat'),  # 18
                flight_data.get('arrival_lon'),  # 19
                flight_data.get('dof_date'),  # 20
                flight_data.get('arrival_actual_date'),  # 21
                flight_data.get('departure_actual_date'),  # 22
                flight_data.get('planned_date'),  # 23
                flight_data.get('atd_time'),  # 24
                flight_data.get('departure_actual_time'),  # 25
                flight_data.get('ata_time'),  # 26
                flight_data.get('arrival_actual_time'),  # 27
                flight_data.get('eet_info'),  # 28 ← УЛУЧШЕННОЕ поле!
                messages.get('shr'),  # 29
                messages.get('dep'),  # 30
                messages.get('arr'),  # 31
                row.get('centr_es_orvd') or row.get('Центр ЕС ОрВД'),  # 32
                flight_data.get('altitude_info'),  # 33
                flight_data.get('altitude_max'),  # 34 ← Максимальная высота (числовое значение)
                flight_data.get('phone_number'),  # 35
                flight_data.get('quantity'),  # 36
                flight_data.get('sts', ''),  # 37 ← Причина особого отношения
            )
            return values  # 37 элементов, filename добавится 38-м

        except Exception as e:
            logging.error(f"Ошибка обработки записи: {e}", exc_info=True)
            return None

    def extract_sid_from_dep_arr(self, text: str) -> Optional[str]:
        """Извлечение SID из сообщений DEP/ARR, где он указан как -SID <число>."""
        if not text:
            return None
        match = re.search(r'-SID\s+(\d+)', text)
        return match.group(1) if match else None

    def parse_shr_message(self, shr_text: str) -> Dict[str, Any]:
        """Парсинг сообщения типа SHR с улучшенной обработкой."""
        if not shr_text:
            return {}

        data = {}

        # Улучшенный поиск SID (учитывает различные форматы)
        sid_patterns = [
            r'SID/(\d{7,15})',
            r'SID\s+(\d{7,15})',
            r'RMK/.*?SID/(\d{7,15})',
            r'RMK/.*?SID\s+(\d{7,15})'
        ]

        for pattern in sid_patterns:
            sid_match = re.search(pattern, shr_text, re.IGNORECASE | re.DOTALL)
            if sid_match:
                data['sid'] = sid_match.group(1)
                break

        # SID из RMK/
        sid_match = re.search(r'RMK/.*?SID/(\d+)', shr_text)
        if sid_match:
            data['sid'] = sid_match.group(1)

      

        # DOF/ дата
        dof_match = re.search(r'DOF/(\d{6})', shr_text)
        if dof_match:
            try:
                yy, mm, dd = map(int, [dof_match.group(1)[i:i + 2] for i in (0, 2, 4)])
                data['planned_date'] = date(2000 + yy, mm, dd)
                data['dof_date'] = data['planned_date']  # Для старого поля
            except ValueError:
                pass

        # TYP/ тип и количество
        typ_match = re.search(r'TYP/(\d*)([A-Z]{3,4})', shr_text)
        if typ_match:
            data['quantity'] = int(typ_match.group(1)) if typ_match.group(1) else 1
            data['aircraft_type_code'] = typ_match.group(2)
            data['aircraft_type_desc'] = self._get_aircraft_desc(typ_match.group(2))
        else:
            # Fallback: пытаемся определить тип из текста remarks
            if 'Беспилотный летательный аппарат' in shr_text or 'дрон' in shr_text.lower():
                data['aircraft_type_code'] = 'BLA'
                data['aircraft_type_desc'] = 'Беспилотный летательный аппарат (дрон)'

        # REG/ регистрационный номер - ИСПОЛЬЗУЕМ УЛУЧШЕННОЕ ИЗВЛЕЧЕНИЕ
        reg_info = self.extract_all_registrations(shr_text)
        if reg_info['registration']:
            data['registration'] = reg_info['registration']
            data['registration_all'] = reg_info['registration_all']
       
        # OPR/ оператор - используем ТОЛЬКО улучшенный метод
        operator_info = self.extract_operator_info(shr_text)
        if operator_info.get('operator'):
            data['operator'] = operator_info['operator']
        if operator_info.get('phone_number'):
            data['phone_number'] = operator_info['phone_number']

        # Телефон уже извлечен через extract_operator_info() выше

        # Номер рейса
        flight_number = self.extract_flight_number(shr_text)
        if flight_number:
            data['flight_number'] = flight_number
        else:
            # Отладочная информация для понимания, почему не извлекается
            if shr_text and len(shr_text) > 10:  # Только для непустых сообщений
                logging.debug(f"Не удалось извлечь номер рейса из SHR: {shr_text[:100]}...")

        # Информация о высоте
        altitude_info = self.extract_altitude_info(shr_text)
        if altitude_info:
            data['altitude_info'] = altitude_info

        # EET (расчетное время полета) - ИСПОЛЬЗУЕМ УЛУЧШЕННУЮ ФУНКЦИЮ
        eet_info = self.extract_eet_info(shr_text)
        if eet_info:
            data['eet_info'] = eet_info

        # STS - причина особого отношения
        data['sts'] = ''
        sts = self.extract_sts(shr_text)
        if sts:
            data['sts'] = sts
        

        # Извлечение координат из SHR - агрессивный поиск
        departure_coords = self.extract_departure_coordinates_aggressive(shr_text)
        if departure_coords:
            data['departure_coords'] = departure_coords['coords']
            data['departure_lat'] = departure_coords['latitude']
            data['departure_lon'] = departure_coords['longitude']
            # Убираем debug логи для производительности

        # Извлечение координат прибытия - агрессивный поиск
        arrival_coords = self.extract_arrival_coordinates_aggressive(shr_text)
        if arrival_coords:
            data['arrival_coords'] = arrival_coords['coords']
            data['arrival_lat'] = arrival_coords['latitude']
            data['arrival_lon'] = arrival_coords['longitude']
            # Убираем debug логи для производительности

        return data

    def clean_operator_text(self, operator_text: str) -> str:
        """Очистка текста оператора от мусора"""
        if not operator_text:
            return ""

        # Удаляем остатки тегов и служебной информации
        cleaned = re.sub(r'\s*(?:REG|TYP|RMK|TEL|PHONE)/.*$', '', operator_text, flags=re.IGNORECASE)

        # Удаляем множественные пробелы
        cleaned = re.sub(r'\s+', ' ', cleaned)

        # Удаляем ведущие/замыкающие символы
        cleaned = cleaned.strip(' ,.-/')

        return cleaned

    def parse_dep_message(self, dep_text: str) -> Dict[str, Any]:
        """Парсинг сообщения типа DEP."""
        if not dep_text:
            return {}

        data = {}
        # -SID
        sid_match = re.search(r'-SID\s+(\d+)', dep_text)
        if sid_match:
            data['sid'] = sid_match.group(1)

        # Извлечение flight_id используя метод extract_sid
        flight_id = self.extract_sid(dep_text)
        if flight_id:
            data['flight_id'] = flight_id

        # -ADD дата
        add_match = re.search(r'-ADD\s+(\d{6})', dep_text)
        if add_match:
            try:
                yy, mm, dd = map(int, [add_match.group(1)[i:i + 2] for i in (0, 2, 4)])
                data['departure_actual_date'] = date(2000 + yy, mm, dd)
            except ValueError:
                pass

        # -ATD время
        atd_match = re.search(r'-ATD\s+(\d{4})', dep_text)
        if atd_match:
            try:
                hh, mm = int(atd_match.group(1)[:2]), int(atd_match.group(1)[2:4])
                data['departure_actual_time'] = time(hh, mm)
                data['atd_time'] = data['departure_actual_time']  # Для старого поля
            except ValueError:
                pass

        # -REG
        reg_match = re.search(r'-REG\s+([^\s]+)', dep_text)
        if reg_match:
            data['registration'] = reg_match.group(1)

        # -ADEPZ координаты
        adepz_match = re.search(r'-ADEPZ\s+([0-9NSWECВСЗЮ]+)', dep_text)
        if adepz_match:
            coords = self.parse_coordinates_universal(adepz_match.group(1))
            if coords:
                data['departure_coords'] = adepz_match.group(1)
                data['departure_lat'] = coords['latitude']
                data['departure_lon'] = coords['longitude']

        # Номер рейса
        flight_number = self.extract_flight_number(dep_text)
        if flight_number:
            data['flight_number'] = flight_number

        # Информация о высоте
        altitude_info = self.extract_altitude_info(dep_text)
        if altitude_info:
            data['altitude_info'] = altitude_info

        # EET (расчетное время полета) - ИСПОЛЬЗУЕМ УЛУЧШЕННУЮ ФУНКЦИЮ
        eet_info = self.extract_eet_info(dep_text)
        if eet_info:
            data['eet_info'] = eet_info

        # # STS - причина особого отношения
        # data['sts'] = ''
        # sts = self.extract_sts(dep_text)
        # if sts:
        #     data['sts'] = sts

        return data

    def parse_arr_message(self, arr_text: str) -> Dict[str, Any]:
        """Парсинг сообщения типа ARR."""
        if not arr_text:
            return {}

        data = {}
        # -SID
        sid_match = re.search(r'-SID\s+(\d+)', arr_text)
        if sid_match:
            data['sid'] = sid_match.group(1)

        # Извлечение flight_id используя метод extract_sid
        flight_id = self.extract_sid(arr_text)
        if flight_id:
            data['flight_id'] = flight_id

        # -ADA дата
        ada_match = re.search(r'-ADA\s+(\d{6})', arr_text)
        if ada_match:
            try:
                yy, mm, dd = map(int, [ada_match.group(1)[i:i + 2] for i in (0, 2, 4)])
                data['arrival_actual_date'] = date(2000 + yy, mm, dd)
            except ValueError:
                pass

        # -ATA время
        ata_match = re.search(r'-ATA\s+(\d{4})', arr_text)
        if ata_match:
            try:
                hh, mm = int(ata_match.group(1)[:2]), int(ata_match.group(1)[2:4])
                data['arrival_actual_time'] = time(hh, mm)
                data['ata_time'] = data['arrival_actual_time']  # Для старого поля
            except ValueError:
                pass

        # -REG
        reg_match = re.search(r'-REG\s+([^\s]+)', arr_text)
        if reg_match:
            data['registration'] = reg_match.group(1)

        # -ADARRZ координаты
        adarrz_match = re.search(r'-ADARRZ\s+([0-9NSWECВСЗЮ]+)', arr_text)
        if adarrz_match:
            coords = self.parse_coordinates_universal(adarrz_match.group(1))
            if coords:
                data['arrival_coords'] = adarrz_match.group(1)
                data['arrival_lat'] = coords['latitude']
                data['arrival_lon'] = coords['longitude']

        # Номер рейса
        flight_number = self.extract_flight_number(arr_text)
        if flight_number:
            data['flight_number'] = flight_number

        # Информация о высоте
        altitude_info = self.extract_altitude_info(arr_text)
        if altitude_info:
            data['altitude_info'] = altitude_info

        # EET (расчетное время полета) - ИСПОЛЬЗУЕМ УЛУЧШЕННУЮ ФУНКЦИЮ
        eet_info = self.extract_eet_info(arr_text)
        if eet_info:
            data['eet_info'] = eet_info

        # STS - причина особого отношения
        data['sts'] = ''
        sts = self.extract_sts(arr_text)
        if sts:
            data['sts'] = sts

        return data

    def _get_aircraft_desc(self, code: str) -> str:
        """Возвращает расшифровку кода типа ВС."""
        descriptions = {
            'BLA': 'Беспилотный летательный аппарат (дрон)',
            'AER': 'Самолет',
            'HEL': 'Вертолет',
            'SHAR': 'Аэростат',
            'GLI': 'Планер',
            'UAV': 'Беспилотный летательный аппарат',
            'GYR': 'Автожир',
            'ULM': 'Сверхлегкий летательный аппарат'
        }
        return descriptions.get(code, 'Неизвестный тип воздушного судна')

    def merge_flight_data(self, shr_data: Dict, dep_data: Dict, arr_data: Dict) -> Dict[str, Any]:
        """Слияние данных с приоритетом фактических (DEP/ARR) над плановыми (SHR)."""
        merged = {}

        # Информация о ВС и операторе (в основном из SHR)
        # Получаем оператора и обрезаем до 255 символов (максимум в БД)
        operator = shr_data.get('operator') or dep_data.get('operator') or arr_data.get('operator')
        if operator and len(operator) > 255:
            logging.warning(f"Оператор обрезан до 255 символов: {operator[:50]}...")
            operator = operator[:255]
        
        merged.update({
            'flight_id': shr_data.get('sid') or dep_data.get('sid') or arr_data.get('sid'),
            'aircraft_type_code': shr_data.get('aircraft_type_code') or dep_data.get(
                'aircraft_type_code') or arr_data.get('aircraft_type_code'),
            'aircraft_type_desc': shr_data.get('aircraft_type_desc') or dep_data.get(
                'aircraft_type_desc') or arr_data.get('aircraft_type_desc'),
            'quantity': shr_data.get('quantity', 1),
            'operator': operator,
            'phone_number': shr_data.get('phone_number') or dep_data.get('phone_number') or arr_data.get(
                'phone_number'),
            'registration': shr_data.get('registration') or dep_data.get('registration') or arr_data.get(
                'registration'),
            'registration_all': shr_data.get('registration_all') or dep_data.get('registration_all') or arr_data.get(
                'registration_all'),
            'planned_date': shr_data.get('planned_date'),
            'dof_date': shr_data.get('dof_date'),
        })

        # Фактические данные из DEP и ARR с fallback на SHR
        merged.update({
            'departure_actual_date': dep_data.get('departure_actual_date') or shr_data.get('planned_date'),
            'departure_actual_time': dep_data.get('departure_actual_time'),
            'atd_time': dep_data.get('atd_time') or shr_data.get('atd_time'),
            'departure_coords': dep_data.get('departure_coords') or shr_data.get('departure_coords'),
            'departure_lat': dep_data.get('departure_lat') or shr_data.get('departure_lat'),
            'departure_lon': dep_data.get('departure_lon') or shr_data.get('departure_lon'),

            'arrival_actual_date': arr_data.get('arrival_actual_date') or shr_data.get('planned_date'),
            'arrival_actual_time': arr_data.get('arrival_actual_time'),
            'ata_time': arr_data.get('ata_time') or shr_data.get('ata_time'),
            'arrival_coords': arr_data.get('arrival_coords') or shr_data.get('arrival_coords'),
            'arrival_lat': arr_data.get('arrival_lat') or shr_data.get('arrival_lat'),
            'arrival_lon': arr_data.get('arrival_lon') or shr_data.get('arrival_lon'),
        })

        # Расчет длительности полета
        dep_dt = None
        arr_dt = None
        if merged['departure_actual_date'] and merged['departure_actual_time']:
            dep_dt = datetime.combine(merged['departure_actual_date'], merged['departure_actual_time'])
        elif merged['dof_date'] and merged['atd_time']:
            dep_dt = datetime.combine(merged['dof_date'], merged['atd_time'])

        if merged['arrival_actual_date'] and merged['arrival_actual_time']:
            arr_dt = datetime.combine(merged['arrival_actual_date'], merged['arrival_actual_time'])
        elif merged['dof_date'] and merged['ata_time']:
            arr_dt = datetime.combine(merged['dof_date'], merged['ata_time'])

        if dep_dt and arr_dt and arr_dt > dep_dt:
            merged['duration_min'] = int((arr_dt - dep_dt).total_seconds() / 60)

        # Дополнительная информация с приоритетом SHR
        altitude_info = shr_data.get('altitude_info') or dep_data.get('altitude_info') or arr_data.get('altitude_info')
        altitude_max = self.extract_altitude_max(altitude_info) if altitude_info else None
        
        merged.update({
            'eet_info': shr_data.get('eet_info') or dep_data.get('eet_info') or arr_data.get('eet_info'),
            'altitude_info': altitude_info,
            'altitude_max': altitude_max,
            'sts': shr_data.get('sts', '') or dep_data.get('sts', '') or arr_data.get('sts', '') or '',
        })

        return merged

    def calculate_time_of_day(self, flight_data):
        """Определяет время суток на основе времени полета с учетом UTC региона"""
        try:
            # Получаем region_id из данных полета
            region_id = flight_data[6] if len(flight_data) > 6 else None  # region_id находится на позиции 6
            
            if not region_id:
                return 'день'
                
            # Получаем UTC для региона
            utc_offset = self.get_region_utc(region_id)
            if utc_offset is None:
                return 'день'
                
            # Ищем любое доступное время из полей времени
            time_fields = [
                flight_data[25],  # atd_time
                flight_data[26],  # departure_actual_time  
                flight_data[28],  # ata_time
                flight_data[29],  # arrival_actual_time
            ]
            
            # Находим первое непустое время
            flight_time = None
            for time_field in time_fields:
                if time_field is not None:
                    flight_time = time_field
                    break
                    
            if not flight_time:
                return 'день'
                
            # Применяем UTC offset
            if isinstance(flight_time, str):
                try:
                    flight_time = datetime.strptime(flight_time, '%H:%M:%S').time()
                except:
                    try:
                        flight_time = datetime.strptime(flight_time, '%H:%M').time()
                    except:
                        return 'день'
                        
            # Определяем время суток в местном часовом поясе
            # UTC offset показывает разность с UTC, поэтому для определения местного времени
            # мы используем время как есть, а UTC offset нужен только для контекста
            local_hour = flight_time.hour
            
            # Определяем время суток
            if 6 <= local_hour < 12:
                return 'утро'
            elif 12 <= local_hour < 18:
                return 'день'
            elif 18 <= local_hour < 24:
                return 'вечер'
            else:  # 0 <= local_hour < 6
                return 'ночь'
                
        except Exception as e:
            logging.error(f"Ошибка при определении времени суток: {e}")
            return 'день'

    def get_region_utc(self, region_id):
        """Получает UTC offset для региона"""
        try:
            # Сначала пытаемся найти в загруженных регионах
            if self.regions:
                for region in self.regions:
                    if region['id'] == region_id:
                        return region.get('utc', 0)
            
            # Если не найдено, загружаем из БД
            if not hasattr(self, 'regions_utc'):
                self.load_regions_utc()
                
            return self.regions_utc.get(region_id, 0)
        except Exception as e:
            logging.error(f"Ошибка при получении UTC для региона {region_id}: {e}")
            return 0
            
    def load_regions_utc(self):
        """Загружает UTC offset для всех регионов"""
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute("SELECT id, UTC FROM regions")
            regions_data = cursor.fetchall()
            
            self.regions_utc = {}
            for region in regions_data:
                self.regions_utc[region['id']] = region['UTC']
                
            logging.info(f"Загружено UTC для {len(self.regions_utc)} регионов")
            return True
        except Exception as e:
            logging.error(f"Ошибка при загрузке UTC регионов: {e}")
            self.regions_utc = {}
            return False

    # ОСТАЛЬНЫЕ СУЩЕСТВУЮЩИЕ ФУНКЦИИ БЕЗ ИЗМЕНЕНИЙ
    def process_unprocessed_files(self, force_reprocess=False):
        """Обрабатывает только те файлы, которые ещё не помечены как 'processed'.
        
        Args:
            force_reprocess (bool): Если True, переобрабатывает все файлы, даже уже обработанные
        """
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                logging.error("Не удалось подключиться к БД для обработки файлов")
                return

        # Загрузка геоданных (если ещё не загружены)
        if not self.regions:
            if not self.load_regions_from_db():
                logging.error("Не удалось загрузить регионы. Прерывание обработки.")
                return

        # Получаем список файлов с FTP
        ftp_files = self.list_ftp_files()
        for ftp_file in ftp_files:
            try:
                if not force_reprocess and self.is_file_processed(ftp_file):
                    continue

                 # Удаляем предыдущие данные для этого файла (если они есть)
                self.delete_flights_by_filename(ftp_file)

                self.mark_file_processed(ftp_file, "process")
                self.current_filename = ftp_file
                self.current_file_progress = 0

                local_path = f"downloads/{ftp_file}"
                os.makedirs("downloads", exist_ok=True)

                if not self.download_ftp_file(f"{settings.FTP_CONFIG['remote_dir']}{ftp_file}", local_path):
                    self.mark_file_processed(ftp_file, "error", "Ошибка загрузки файла")
                    continue

                data = self.load_data_from_excel(local_path)
                if data is None:
                    self.mark_file_processed(ftp_file, "error", "Ошибка загрузки данных")
                    continue

                self.process_flights(data, limit=None, batch_size=10000)
                self.mark_file_processed(ftp_file, "processed")
                self.update_progress(1, 1, f"Файл обработан")
                self.remove_duplicates_by_uniq_str()

            except Exception as e:
                error_msg = f"Ошибка обработки файла {ftp_file}: {e}"
                logging.error(error_msg)
                self.mark_file_processed(ftp_file, "error", error_msg)

    def connect_to_db(self):
        """Установка соединения с базой данных"""
        try:
            self.connection = mysql.connector.connect(**settings.DB_CONFIG)
            logging.info("Успешное подключение к базе данных")
            return True
        except Error as e:
            logging.error(f"Ошибка подключения к MySQL: {e}")
            return False

    def create_log_table(self):
        """Создание таблицы LOG для записи сообщений"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS log (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    message TEXT,
                    procent DECIMAL(5,2)
                )
            """)
            self.connection.commit()
            logging.info("Таблица log создана или уже существует")
        except Exception as e:
            logging.error(f"Ошибка создания таблицы log: {e}")
        finally:
            if cursor:
                cursor.close()

    def write_log(self, message, procent=None):
        """Запись сообщения в таблицу LOG"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("INSERT INTO log (message, procent) VALUES (%s, %s)",
                           (message, procent))
            self.connection.commit()
        except Exception as e:
            logging.error(f"Ошибка записи в лог: {e}")
        finally:
            if cursor:
                cursor.close()

    def update_progress(self, current, total, stage="processing"):
        """Обновление прогресса выполнения"""
        if total == 0:
            return

        progress_percent = round((current / total) * 100, 2)
        self.current_file_progress = progress_percent

        # Записываем в лог
        log_message = f"{stage}: ({progress_percent}%)"
        self.write_log(log_message, progress_percent)

        # Обновляем поле procent в processed_files
        if self.current_filename:
            try:
                cursor = self.connection.cursor()
                cursor.execute("""
                    UPDATE processed_files 
                    SET procent = %s 
                    WHERE filename = %s
                """, (progress_percent, self.current_filename))
                self.connection.commit()
            except Exception as e:
                logging.error(f"Ошибка обновления прогресса: {e}")
            finally:
                if cursor:
                    cursor.close()

    def list_ftp_files(self):
        """Получение списка файлов с FTP"""
        try:
            ftp = FTP(settings.FTP_CONFIG['host'])
            ftp.login(settings.FTP_CONFIG['username'], settings.FTP_CONFIG['password'])
            ftp.cwd(settings.FTP_CONFIG['remote_dir'])
            files = ftp.nlst()
            ftp.quit()
            xlsx_files = [f for f in files if f.endswith('.xlsx')]
            logging.info(f"Найдено не обработанных файлов: {len(xlsx_files)}")
            self.update_progress(0, 1, f"Найдено не обработанных файлов: {len(xlsx_files)}")
            return xlsx_files
        except Exception as e:
            logging.error(f"Ошибка получения списка файлов с FTP: {e}")
            return []

    def download_ftp_file(self, remote_path, local_path):
        """Загрузка файла с FTP"""
        try:
            ftp = FTP(settings.FTP_CONFIG['host'])
            ftp.login(settings.FTP_CONFIG['username'], settings.FTP_CONFIG['password'])

            with open(local_path, 'wb') as f:
                ftp.retrbinary(f"RETR {remote_path}", f.write)

            ftp.quit()
            logging.info(f"Файл успешно загружен для обработки: {remote_path} -> {local_path}")
            self.update_progress(0, 1, f"Файл успешно загружен для обработки: {local_path}")
            return True
        except Exception as e:
            logging.error(f"Ошибка загрузки файла с FTP {remote_path}: {e}")
            return False

    def create_processed_files_table(self):
        """Создание таблицы для отслеживания обработанных файлов"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS processed_files (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    filename VARCHAR(255) NOT NULL,
                    processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('processed', 'error') NOT NULL,
                    error_message TEXT,
                    procent DECIMAL(5,2) DEFAULT 0.00,
                    UNIQUE KEY unique_filename (filename)
                ) COMMENT='Таблица для учета обработанных файлов'
            """)
            self.connection.commit()
            logging.info("Таблица processed_files создана или уже существует")
        except Exception as e:
            logging.error(f"Ошибка создания таблица processed_files: {e}")
        finally:
            if cursor:
                cursor.close()

    def delete_flights_by_filename(self, filename):
        """Удаление всех записей из processed_flights по имени файла"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("DELETE FROM processed_flights WHERE filename = %s", (filename,))
            deleted_count = cursor.rowcount
            self.connection.commit()
            if deleted_count > 0:
                logging.info(f"Удалено {deleted_count} записей для файла {filename}")
            return deleted_count
        except Exception as e:
            logging.error(f"Ошибка удаления записей для файла {filename}: {e}")
            return 0
        finally:
            if cursor:
                cursor.close()

    def mark_file_processed(self, filename, status="processed", error_message=None):
        """Отметка файла как обработанного"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                INSERT INTO processed_files (filename, status, error_message, procent)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                status = VALUES(status),
                error_message = VALUES(error_message),
                processed_at = CURRENT_TIMESTAMP,
                procent = VALUES(procent)
            """, (filename, status, error_message, self.current_file_progress))
            self.connection.commit()
        except Exception as e:
            logging.error(f"Ошибка отметки файла как обработанного: {e}")
        finally:
            if cursor:
                cursor.close()

    def is_file_processed(self, filename):
        """Проверка, был ли файл уже обработан"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT status FROM processed_files WHERE filename = %s", (filename,))
            result = cursor.fetchone()
            return result is not None and result[0] == "processed"
        except Exception as e:
            logging.error(f"Ошибка проверки статуса файла: {e}")
            return False
        finally:
            if cursor:
                cursor.close()

    def detect_sheet_structure(self, sheet_name, df):
        """Определение структуры листа Excel"""
        structure = {
            'has_shr': False,
            'has_dep': False,
            'has_arr': False,
            'date_column': None,
            'shr_column': None,
            'dep_column': None,
            'arr_column': None,
            'center_column': None
        }

        for col in df.columns:
            col_lower = str(col).lower()
            if 'shr' in col_lower:
                structure['has_shr'] = True
                structure['shr_column'] = col
            elif 'dep' in col_lower and 'arr' not in col_lower:
                structure['has_dep'] = True
                structure['dep_column'] = col
            elif 'arr' in col_lower:
                structure['has_arr'] = True
                structure['arr_column'] = col
            elif 'центр' in col_lower or 'center' in col_lower:
                structure['center_column'] = col
            elif 'дата' in col_lower or 'date' in col_lower:
                structure['date_column'] = col

        logging.info(f"Структура листа '{sheet_name}': {structure}")

        return structure

    def load_data_from_excel(self, file_path):
        """Загрузка данных из Excel файла с поддержкой нескольких листов"""
        try:
            # Загружаем книгу Excel
            wb = load_workbook(file_path)
            all_data = []
            f = 0
            for sheet_name in wb.sheetnames:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                f = f + 1
                if df.empty:
                    continue

                # Определяем структуру листа
                structure = self.detect_sheet_structure(sheet_name, df)
                self.update_progress(f, len(wb.sheetnames) * 10, f"Обработка листа '{sheet_name}'")
                self.sheet_structures[sheet_name] = structure

                # Добавляем информацию о листе в каждую запись
                for idx, row in df.iterrows():
                    row_dict = row.to_dict()
                    row_dict['sheet_name'] = sheet_name
                    row_dict['sheet_structure'] = structure
                    row_dict['source_id'] = idx  # Используем индекс как временный ID
                    all_data.append(row_dict)

            logging.info(f"Загружено {len(all_data)} записей из {len(wb.sheetnames)} листов Excel файла")
            return all_data

        except Exception as e:
            error_msg = f"Ошибка загрузки данных из Excel {file_path}: {e}"
            logging.error(error_msg)
            return None

    def extract_messages(self, row):
        """Извлечение текстов SHR, DEP, ARR из строки"""
        structure = row.get('sheet_structure', {})

        messages = {
            'shr': None,
            'dep': None,
            'arr': None
        }

        # Извлекаем SHR
        if structure.get('has_shr') and structure.get('shr_column'):
            shr_value = row.get(structure['shr_column'])
            # Преобразуем float в строку, если нужно
            if isinstance(shr_value, float) and not pd.isna(shr_value):
                shr_value = str(int(shr_value)) if shr_value.is_integer() else str(shr_value)
            if shr_value and not pd.isna(shr_value):
                # Очищаем от переносов строк перед обработкой
                messages['shr'] = self.clean_line_breaks(str(shr_value))

        # Если нет отдельной колонки SHR, ищем в других колонках
        if not messages['shr']:
            for col, value in row.items():
                if isinstance(value, str) and value.startswith('(SHR-'):
                    # Очищаем от переносов строк перед обработкой
                    messages['shr'] = self.clean_line_breaks(value)
                    break

        # Извлекаем DEP
        if structure.get('has_dep') and structure.get('dep_column'):
            dep_value = row.get(structure['dep_column'])
            # Преобразуем float в строку, если нужно
            if isinstance(dep_value, float) and not pd.isna(dep_value):
                dep_value = str(int(dep_value)) if dep_value.is_integer() else str(dep_value)
            if dep_value and not pd.isna(dep_value):
                # Очищаем от переносов строк перед обработкой
                messages['dep'] = self.clean_line_breaks(str(dep_value))

        # Извлекаем ARR
        if structure.get('has_arr') and structure.get('arr_column'):
            arr_value = row.get(structure['arr_column'])
            # Преобразуем float в строку, если нужно
            if isinstance(arr_value, float) and not pd.isna(arr_value):
                arr_value = str(int(arr_value)) if arr_value.is_integer() else str(arr_value)
            if arr_value and not pd.isna(arr_value):
                # Очищаем от переносов строк перед обработкой
                messages['arr'] = self.clean_line_breaks(str(arr_value))

        return messages

    def load_regions_from_db(self):
        """Загрузка геоданных регионов из базы данных с улучшенной обработкой"""
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute("SELECT id, name, polygon, UTC FROM regions")
            regions_data = cursor.fetchall()

            for region in regions_data:
                try:
                    # Пытаемся распарсить полигон в разных форматах
                    polygon_data = None

                    # Попытка 1: как JSON
                    try:
                        polygon_data = json.loads(region['polygon'])
                    except:
                        pass

                    # Попытка 2: как Python literal
                    if not polygon_data:
                        try:
                            polygon_data = ast.literal_eval(region['polygon'])
                        except:
                            pass

                    # Попытка 3: как текст с координатами
                    if not polygon_data:
                        try:
                            # Формат: "55.7558,37.6176;55.7558,37.6176;..."
                            coords = []
                            for coord_pair in region['polygon'].split(';'):
                                if coord_pair.strip():
                                    lat, lon = coord_pair.split(',')
                                    coords.append((float(lon), float(lat)))
                            polygon_data = [coords]
                        except:
                            logging.error(f"Не удалось распарсить полигон для региона {region['name']}")
                            continue

                    if polygon_data and len(polygon_data) > 0:
                        # Преобразуем координаты в правильный порядок (долгота, широта)
                        main_polygon = polygon_data[0] if isinstance(polygon_data, list) else polygon_data

                        # Убедимся, что координаты в правильном порядке
                        formatted_coords = []
                        for coord in main_polygon:
                            if len(coord) >= 2:
                                # Если первое значение больше 90, это вероятно долгота
                                if abs(coord[0]) > 90 and abs(coord[1]) <= 90:
                                    formatted_coords.append((coord[0], coord[1]))  # долгота, широта
                                else:
                                    formatted_coords.append((coord[1], coord[0]))  # меняем местами

                        if len(formatted_coords) < 3:
                            logging.warning(
                                f"Полигон региона {region['name']} имеет недостаточно точек: {len(formatted_coords)}")
                            continue

                        polygon = Polygon(formatted_coords)
                        if not polygon.is_valid:
                            # Пытаемся исправить невалидный полигон
                            polygon = polygon.buffer(0)

                        prepped_polygon = prep(polygon)

                        self.regions.append({
                            'id': region['id'],
                            'name': region['name'],
                            'utc': region.get('UTC', 0),
                            'polygon': polygon,
                            'prepped_polygon': prepped_polygon
                        })
                except Exception as e:
                    logging.error(f"Ошибка обработки полигона для региона {region['name']}: {e}")
                    continue

            # Создаем пространственный индекс для быстрого поиска
            polygons = [region['polygon'] for region in self.regions]
            self.region_tree = STRtree(polygons)

            logging.info(f"Загружено {len(self.regions)} регионов из базы данных")
            return True

        except Exception as e:
            logging.error(f"Ошибка загрузки регионов из базы данных: {e}")
            return False
        finally:
            if cursor:
                cursor.close()

    def parse_coordinates_universal(self, coord_str):
        """Универсальный парсинг координат с улучшенной обработкой форматов"""
        if not coord_str or not isinstance(coord_str, str):
            return None

        coord_str = coord_str.strip().upper()

        # Удаляем лишние символы (скобки, пробелы и т.д.), но оставляем русские буквы
        coord_str = re.sub(r'[^\dNSWECВСЗЮ]', '', coord_str)

        # Заменяем русские буквы на английские
        coord_str = coord_str.replace('С', 'N').replace('В', 'E').replace('З', 'W').replace('Ю', 'S')

        # Универсальный паттерн для любых координат: цифры + N/S + цифры + E/W
        universal_match = re.match(r'(\d+)([NS])(\d+)([EW])', coord_str)
        if not universal_match:
            # Проверяем, может быть это неполная координата (только широта или только долгота)
            if re.match(r'^\d+$', coord_str):
                return None
            return None
        
        lat_str, lat_dir, lon_str, lon_dir = universal_match.groups()
        # Убираем debug логи для производительности

        # Парсим широту - гибкая логика для любого количества цифр
        lat_degrees, lat_minutes, lat_seconds = 0, 0, 0
        
        if len(lat_str) >= 6:
            # DDMMSS - градусы, минуты, секунды
            lat_degrees = int(lat_str[0:2])
            lat_minutes = int(lat_str[2:4])
            lat_seconds = int(lat_str[4:6])
        elif len(lat_str) == 5:
            # DDMMM - градусы и минуты (5 цифр)
            lat_degrees = int(lat_str[0:2])
            lat_minutes = int(lat_str[2:5])
        elif len(lat_str) == 4:
            # DDMM - градусы и минуты (4 цифры)
            lat_degrees = int(lat_str[0:2])
            lat_minutes = int(lat_str[2:4])
        elif len(lat_str) == 3:
            # DDM - градусы и минуты (3 цифры)
            lat_degrees = int(lat_str[0:2])
            lat_minutes = int(lat_str[2:3])
        elif len(lat_str) == 2:
            # DD - только градусы
            lat_degrees = int(lat_str)
        else:
            # Для других случаев пытаемся разбить пополам
            if len(lat_str) >= 4:
                mid = len(lat_str) // 2
                lat_degrees = int(lat_str[:mid])
                lat_minutes = int(lat_str[mid:])
            else:
                lat_degrees = int(lat_str)

        # Парсим долготу - гибкая логика для любого количества цифр
        lon_degrees, lon_minutes, lon_seconds = 0, 0, 0
        
        if len(lon_str) >= 7:
            # DDDMMSS - градусы, минуты, секунды
            lon_degrees = int(lon_str[0:3])
            lon_minutes = int(lon_str[3:5])
            lon_seconds = int(lon_str[5:7])
        elif len(lon_str) == 6:
            # DDDMMS - градусы, минуты, секунды (6 цифр)
            lon_degrees = int(lon_str[0:3])
            lon_minutes = int(lon_str[3:5])
            lon_seconds = int(lon_str[5:6])
        elif len(lon_str) == 5:
            # Для 5 цифр: определяем формат по разумности значений
            if lon_str[0] == '0':
                # 03322 -> 033 градуса, 22 минуты
                if int(lon_str[0:3]) <= 180:  # Проверяем разумность градусов
                    lon_degrees = int(lon_str[0:3])
                    lon_minutes = int(lon_str[3:5])
                else:
                    # 2+3: первые 2 цифры - градусы, остальные 3 - минуты
                    lon_degrees = int(lon_str[0:2])
                    lon_minutes = int(lon_str[2:5])
            else:
                # Для долгот, не начинающихся с 0: проверяем разумность
                if int(lon_str[0:3]) > 180:
                    # 12734 -> 12 градусов, 734 минуты (неправильно!)
                    # Лучше 127 градусов, 34 минуты
                    if int(lon_str[0:3]) <= 180:
                        lon_degrees = int(lon_str[0:3])
                        lon_minutes = int(lon_str[3:5])
                    else:
                        # 2+3: первые 2 цифры - градусы, остальные 3 - минуты
                        lon_degrees = int(lon_str[0:2])
                        lon_minutes = int(lon_str[2:5])
                else:
                    # 12734 -> 127 градусов, 34 минуты (правильно!)
                    lon_degrees = int(lon_str[0:3])
                    lon_minutes = int(lon_str[3:5])
        elif len(lon_str) == 4:
            # DDDM или DDMM - определяем по первой цифре
            if int(lon_str[0:3]) > 100:
                # DDDM - 3 цифры градусы, 1 цифра минуты
                lon_degrees = int(lon_str[0:3])
                lon_minutes = int(lon_str[3:4])
            else:
                # DDMM - 2 цифры градусы, 2 цифры минуты
                lon_degrees = int(lon_str[0:2])
                lon_minutes = int(lon_str[2:4])
        elif len(lon_str) == 3:
            # DDD - только градусы
            lon_degrees = int(lon_str)
        elif len(lon_str) == 2:
            # DD - только градусы
            lon_degrees = int(lon_str)
        else:
            # Для других случаев пытаемся разбить пополам
            if len(lon_str) >= 4:
                mid = len(lon_str) // 2
                lon_degrees = int(lon_str[:mid])
                lon_minutes = int(lon_str[mid:])
            else:
                lon_degrees = int(lon_str)

        # Преобразуем в десятичные градусы
        latitude = lat_degrees + (lat_minutes / 60) + (lat_seconds / 3600)
        if lat_dir == 'S':
            latitude = -latitude

        longitude = lon_degrees + (lon_minutes / 60) + (lon_seconds / 3600)
        if lon_dir == 'W':
            longitude = -longitude

        return {
            'original': coord_str,
            'latitude': round(latitude, 6),
            'longitude': round(longitude, 6)
        }

    def extract_coordinates_from_zona(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из ZONA секций"""
        if not shr_text:
            return None

        # Паттерны для поиска координат в ZONA секциях
        zona_patterns = [
            # /ZONA ZONA R001 5456N01959E/
            r'/ZONA\s+ZONA\s+R\d+\s+([0-9NSWECВСЗЮ]+)/',
            # /ZONA R001 5456N01959E/
            r'/ZONA\s+R\d+\s+([0-9NSWECВСЗЮ]+)/',
            # ZONA R001 5456N01959E
            r'ZONA\s+R\d+\s+([0-9NSWECВСЗЮ]+)',
            # Более общий паттерн для ZONA
            r'ZONA[^/]*?([0-9]{4,6}[NS][0-9]{4,7}[EW])',
        ]

        for pattern in zona_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                # Убираем debug логи для производительности
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
                # Убираем warning логи для производительности

        # Если не нашли одиночные координаты, ищем полигон ZONA
        # /ZONA 5344N05104E 5341N05109E 5317N05034E 5316N05034E 5316N05029E 5338N05052E 5344N05104E/
        polygon_pattern = r'/ZONA\s+([0-9NSWECВСЗЮ\s]+)/'
        polygon_match = re.search(polygon_pattern, shr_text, re.IGNORECASE)
        if polygon_match:
            coords_text = polygon_match.group(1)
            logging.debug(f"Найден полигон ZONA: {coords_text}")
            
            # Извлекаем все координаты из полигона
            coord_matches = re.findall(r'([0-9]{4,6}[NS][0-9]{4,7}[EW])', coords_text)
            if coord_matches:
                # Берем первую координату как departure
                coord_str = coord_matches[0]
                # Убираем debug логи для производительности
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
                # Убираем warning логи для производительности

        return None

    def extract_coordinates_from_rmk(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из RMK секций"""
        if not shr_text:
            return None

        # Паттерны для поиска координат в RMK секциях
        rmk_patterns = [
            # С ЦЕНТРОМ 5456N01959E
            r'С\s+ЦЕНТРОМ\s+([0-9NSWECВСЗЮ]+)',
            # ЦЕНТРОМ 5456N01959E
            r'ЦЕНТРОМ\s+([0-9NSWECВСЗЮ]+)',
            # CENTER 5456N01959E
            r'CENTER\s+([0-9NSWECВСЗЮ]+)',
            # Более общий паттерн для координат в RMK
            r'RMK/[^/]*?([0-9]{4,6}[NS][0-9]{4,7}[EW])',
            # Поиск координат в формате 543607N0214703E
            r'([0-9]{6,7}[NS][0-9]{6,7}[EW])',
        ]

        for pattern in rmk_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                # Убираем debug логи для производительности
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
                # Убираем warning логи для производительности

        return None

    def extract_departure_coordinates_aggressive(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Агрессивный поиск координат отправления всеми возможными способами с оценкой качества"""
        if not shr_text:
            return None

        # Список всех возможных методов поиска координат с их приоритетами
        extraction_methods = [
            (self._extract_dep_coordinates, "DEP/ секция", 10),
            (self._extract_zona_coordinates, "ZONA секция", 8),
            (self._extract_rmk_coordinates, "RMK секция", 6),
            (self._extract_eet_coordinates, "EET секция", 7),
            (self._extract_operator_coordinates, "OPR секция", 5),
            (self._extract_any_coordinates_in_text, "Общий поиск", 3),
            (self._extract_fallback_coordinates, "Fallback методы", 1),
        ]

        best_result = None
        best_quality = 0

        for method, method_name, base_quality in extraction_methods:
            try:
                coords = method(shr_text)
                if coords:
                    # Упрощенная оценка качества для скорости
                    quality = base_quality
                    if len(coords['coords']) >= 12:
                        quality += 2
                    elif len(coords['coords']) >= 10:
                        quality += 1
                    
                    if quality > best_quality:
                        best_quality = quality
                        best_result = coords
                        # Если нашли высококачественный результат, можно прервать поиск
                        if quality >= 15:  # DEP/DEST с хорошим качеством
                            break
                    
            except Exception as e:
                # Убираем warning логи для производительности
                continue

        return best_result

    def _extract_dep_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение DEP/ координат"""
        dep_patterns = [
            r'DEP/\s*([0-9NSWECВСЗЮ]+)',  # DEP/5331N05050E
            r'DEP\s+([0-9NSWECВСЗЮ]+)',   # DEP 5331N05050E
            r'DEP:([0-9NSWECВСЗЮ]+)',     # DEP:5331N05050E
            r'DEP\/([0-9NSWE]+)', 
        ]
        
        for pattern in dep_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                # Очищаем координаты от лишних символов
                coord_str = re.sub(r'[^\dNSWECВСЗЮ]', '', coord_str)
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        return None

    def _extract_zona_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из ZONA секций с расширенным поиском"""
        if not shr_text:
            return None

        # Расширенные паттерны для поиска координат в ZONA секциях
        zona_patterns = [
            # /ZONA ZONA R001 5456N01959E/
            r'/ZONA\s+ZONA\s+R\d+\s+([0-9NSWECВСЗЮ]+)/',
            # /ZONA R001 5456N01959E/
            r'/ZONA\s+R\d+\s+([0-9NSWECВСЗЮ]+)/',
            # /ZONA 5456N01959E/
            r'/ZONA\s+([0-9NSWECВСЗЮ]+)/',
            # /ZONA R1,9 4652N14245E/
            r'/ZONA\s+R[0-9,.]+\s+([0-9NSWECВСЗЮ]+)/',
            # /ZONA MR01257/ (без координат, но может быть в тексте)
        ]

        for pattern in zona_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }

        # Если не нашли одиночные координаты, ищем полигон ZONA
        polygon_pattern = r'/ZONA\s+([0-9NSWECВСЗЮ\s]+)/'
        polygon_match = re.search(polygon_pattern, shr_text, re.IGNORECASE)
        if polygon_match:
            coords_text = polygon_match.group(1)
            # Извлекаем все координаты из полигона
            coord_matches = re.findall(r'([0-9]{4,8}[NS][0-9]{4,8}[EW])', coords_text)
            if coord_matches:
                # Берем первую координату как departure
                coord_str = coord_matches[0]
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }

        return None

    def _extract_rmk_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из RMK секций с расширенным поиском"""
        if not shr_text:
            return None

        # Расширенные паттерны для поиска координат в RMK секциях
        rmk_patterns = [
            # С ЦЕНТРОМ 5456N01959E
            r'С\s+ЦЕНТРОМ\s+([0-9NSWECВСЗЮ]+)',
            # ЦЕНТРОМ 5456N01959E
            r'ЦЕНТРОМ\s+([0-9NSWECВСЗЮ]+)',
            # ЦЕНТР 5456N01959E
            r'ЦЕНТР\s+([0-9NSWECВСЗЮ]+)',
            # ГЕОТО4КА 584809N0332228E
            r'ГЕОТО[0-9]КА\s+([0-9NSWECВСЗЮ]+)',
            # ВЗЛЕТ И ПОСАДКА ГЕОТО4КА 584809N0332228E
            r'ВЗЛЕТ\s+И\s+ПОСАДКА\s+ГЕОТО[0-9]КА\s+([0-9NSWECВСЗЮ]+)',
            # ОКРУЖНОСТЬ РАДИУС 6КМ ЦЕНТР 591715N0351341E
            r'ОКРУЖНОСТЬ\s+РАДИУС\s+[0-9]+КМ\s+ЦЕНТР\s+([0-9NSWECВСЗЮ]+)',
            # МНОГОУГОЛЬНИК 583531N0331359E
            r'МНОГОУГОЛЬНИК\s+([0-9NSWECВСЗЮ]+)',
            # ПОЛОСА 585852N0335310E
            r'ПОЛОСА\s+([0-9NSWECВСЗЮ]+)',
        ]

        for pattern in rmk_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        return None

    def _extract_any_coordinates_in_text(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Поиск любых координат в тексте с улучшенными паттернами"""
        # Расширенные паттерны для поиска координат
        coord_patterns = [
            # Стандартные форматы
            r'([0-9]{4,6}[NS][0-9]{4,7}[EW])',  # 5018N12734E, 584809N0332228E
            r'([0-9]{3,6}[NS][0-9]{3,7}[EW])',  # 4652N14245E, 5848N03322E
            # Форматы с дополнительными цифрами
            r'([0-9]{4,8}[NS][0-9]{4,8}[EW])',  # Более широкий диапазон
        ]
        
        # Сначала ищем в DEP/DEST секциях (приоритет)
        dep_dest_pattern = r'(?:DEP|DEST)/\s*([0-9NSWECВСЗЮ]+)'
        dep_dest_matches = re.findall(dep_dest_pattern, shr_text, re.IGNORECASE)
        for coord_str in dep_dest_matches:
            coord_str = re.sub(r'[^\dNSWECВСЗЮ]', '', coord_str)
            coords = self.parse_coordinates_universal(coord_str)
            if coords and -90 <= coords['latitude'] <= 90 and -180 <= coords['longitude'] <= 180:
                return {
                    'coords': coord_str,
                    'latitude': coords['latitude'],
                    'longitude': coords['longitude']
                }
        
        # Затем ищем в остальном тексте
        for pattern in coord_patterns:
            matches = re.findall(pattern, shr_text, re.IGNORECASE)
            for coord_str in matches:
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    # Проверяем, что координаты разумные (в пределах Земли)
                    if -90 <= coords['latitude'] <= 90 and -180 <= coords['longitude'] <= 180:
                        return {
                            'coords': coord_str,
                            'latitude': coords['latitude'],
                            'longitude': coords['longitude']
                        }
        return None

    def extract_arrival_coordinates_aggressive(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Агрессивный поиск координат прибытия всеми возможными способами с оценкой качества"""
        if not shr_text:
            return None

        # Список всех возможных методов поиска координат прибытия с их приоритетами
        extraction_methods = [
            (self._extract_dest_coordinates, "DEST/ секция", 10),
            (self._extract_zona_coordinates, "ZONA секция", 8),
            (self._extract_rmk_coordinates, "RMK секция", 6),
            (self._extract_eet_coordinates, "EET секция", 7),
            (self._extract_operator_coordinates, "OPR секция", 5),
            (self._extract_any_coordinates_in_text, "Общий поиск", 3),
            (self._extract_fallback_coordinates, "Fallback методы", 1),
        ]

        best_result = None
        best_quality = 0

        for method, method_name, base_quality in extraction_methods:
            try:
                coords = method(shr_text)
                if coords:
                    # Упрощенная оценка качества для скорости
                    quality = base_quality
                    if len(coords['coords']) >= 12:
                        quality += 2
                    elif len(coords['coords']) >= 10:
                        quality += 1
                    
                    if quality > best_quality:
                        best_quality = quality
                        best_result = coords
                        # Если нашли высококачественный результат, можно прервать поиск
                        if quality >= 15:  # DEP/DEST с хорошим качеством
                            break
                    
            except Exception as e:
                # Убираем warning логи для производительности
                continue

        return best_result

    def _extract_dest_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение DEST/ координат"""
        dest_patterns = [
            r'DEST/\s*([0-9NSWECВСЗЮ]+)',  # DEST/5331N05050E
            r'DEST\s+([0-9NSWECВСЗЮ]+)',   # DEST 5331N05050E
            r'DEST:([0-9NSWECВСЗЮ]+)',     # DEST:5331N05050E
        ]
        
        for pattern in dest_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(1)
                # Очищаем координаты от лишних символов
                coord_str = re.sub(r'[^\dNSWECВСЗЮ]', '', coord_str)
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        return None

    # Удалена функция _assess_coordinate_quality для оптимизации производительности

    def _extract_eet_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из EET секций"""
        if not shr_text:
            return None
            
        # Паттерны для поиска координат в EET секциях
        eet_patterns = [
            r'EET/([A-Z]{4})(\d{4}[NS]\d{5}[EW])',  # EET/UHHH0001 5022N12800E
            r'EET/([A-Z]{4})\s*(\d{4}[NS]\d{5}[EW])',  # EET/UHHH 5022N12800E
            r'EET/\s*(\d{4}[NS]\d{5}[EW])',  # EET/5022N12800E
        ]
        
        for pattern in eet_patterns:
            match = re.search(pattern, shr_text, re.IGNORECASE)
            if match:
                coord_str = match.group(-1)  # Берем последнюю группу (координаты)
                # Убираем debug логи для производительности
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        return None

    def _extract_operator_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Извлечение координат из OPR секций (в тексте оператора)"""
        if not shr_text:
            return None
            
        # Ищем координаты в тексте оператора
        opr_match = re.search(r'OPR/([^/]+)', shr_text)
        if opr_match:
            opr_text = opr_match.group(1)
            # Ищем координаты в тексте оператора
            coord_patterns = [
                r'(\d{4}[NS]\d{5}[EW])',  # 5022N12800E
                r'(\d{3}[NS]\d{4}[EW])',  # 502N1280E
            ]
            
            for pattern in coord_patterns:
                matches = re.findall(pattern, opr_text, re.IGNORECASE)
                for coord_str in matches:
                    coords = self.parse_coordinates_universal(coord_str)
                    if coords:
                        # Убираем debug логи для производительности
                        return {
                            'coords': coord_str,
                            'latitude': coords['latitude'],
                            'longitude': coords['longitude']
                        }
        return None

    def _extract_fallback_coordinates(self, shr_text: str) -> Optional[Dict[str, Any]]:
        """Fallback методы для поиска координат"""
        if not shr_text:
            return None
            
        # Метод 1: Поиск координат в скобках
        bracket_patterns = [
            r'\(([0-9NSWECВСЗЮ]+)\)',  # (5022N12800E)
            r'\[([0-9NSWECВСЗЮ]+)\]',  # [5022N12800E]
        ]
        
        for pattern in bracket_patterns:
            matches = re.findall(pattern, shr_text, re.IGNORECASE)
            for coord_str in matches:
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    # Убираем debug логи для производительности
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        
        # Метод 2: Поиск координат после слешей
        slash_patterns = [
            r'/([0-9NSWECВСЗЮ]+)/',  # /5022N12800E/
            r'/([0-9NSWECВСЗЮ]+)\s',  # /5022N12800E 
        ]
        
        for pattern in slash_patterns:
            matches = re.findall(pattern, shr_text, re.IGNORECASE)
            for coord_str in matches:
                coords = self.parse_coordinates_universal(coord_str)
                if coords:
                    # Убираем debug логи для производительности
                    return {
                        'coords': coord_str,
                        'latitude': coords['latitude'],
                        'longitude': coords['longitude']
                    }
        
        # Метод 3: Поиск координат в конце строк
        line_end_patterns = [
            r'([0-9NSWECВСЗЮ]+)\s*$',  # 5022N12800E в конце строки
        ]
        
        lines = shr_text.split('\n')
        for line in lines:
            for pattern in line_end_patterns:
                match = re.search(pattern, line.strip(), re.IGNORECASE)
                if match:
                    coord_str = match.group(1)
                    coords = self.parse_coordinates_universal(coord_str)
                    if coords:
                        # Убираем debug логи для производительности
                        return {
                            'coords': coord_str,
                            'latitude': coords['latitude'],
                            'longitude': coords['longitude']
                        }
        
        return None

    def find_region(self, lat: float, lon: float) -> Optional[str]:
        """Поиск региона по координатам с использованием пространственного индекса"""
        if lat is None or lon is None or not self.regions or not self.region_tree:
            return None

        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            return None
        point = Point(lon, lat)

        try:
            # Используем пространственный индекс для быстрого поиска
            possible_indices = self.region_tree.query(point)

            for idx in possible_indices:
                region = self.regions[idx]
                if region['prepped_polygon'].contains(point):
                    return [region['id'], region['name']]

        except Exception as e:
            logging.error(f"Ошибка при поиске региона: {e}")

        return None

    def parse_quantity(self, typ_str: str) -> int:
        """Извлечение количества БПЛА из строки TYP"""
        if not typ_str or not isinstance(typ_str, str):
            return 1

        # Ищем паттерн: число перед BLA
        match = re.search(r'(\d+)BLA', typ_str, re.IGNORECASE)
        if match:
            return int(match.group(1))
        else:
            # Если нет числа, то по умолчанию 1
            return 1

    def extract_sid(self, text: str) -> Optional[str]:
        """Извлечение SID с валидацией как в PHP коде"""
        if not text or not isinstance(text, str):
            return None

        # Более строгое регулярное выражение
        match = re.search(r'SID\/(\d{7,15})\b', text)
        if match:
            sid = match.group(1)

            # Проверяем, что SID состоит только из цифр и имеет разумную длину
            if sid.isdigit() and len(sid) >= 7 and len(sid) <= 15:
                return sid

        return None

    def extract_aircraft_type(self, text: str) -> Dict[str, Any]:
        """Извлечение типа воздушного судна с информацией как в PHP коде"""
        type_codes = {
            'BLA': 'Беспилотный летательный аппарат (дрон)',
            'AER': 'Самолет',
            'HEL': 'Вертолет',
            'SHAR': 'Аэростат',
            'GLI': 'Планер',
            'UAV': 'Беспилотный летательный аппарат',
            'GYR': 'Автожир',
            'ULM': 'Сверхлегкий летательный аппарат'
        }

        if not text or not isinstance(text, str):
            return {'code': 'UNKNOWN', 'quantity': 1}

        # Ищем паттерн TYP/ с возможными цифрами перед типом
        match = re.search(r'TYP\/(\d*)([A-Z]{3,4})', text)
        if match:
            quantity = int(match.group(1)) if match.group(1) else 1
            type_code = match.group(2)
            type_description = type_codes.get(type_code, 'Неизвестный тип воздушного судна')

            return {
                'code': type_code,
                'description': type_description,
                'quantity': quantity
            }

        return {'code': 'UNKNOWN', 'quantity': 1}

    def extract_flight_data(self, messages: Dict[str, str]) -> Dict[str, Any]:
        """Извлечение данных о полете из всех источников с улучшенным парсингом"""
        result = {
            'start_coordinates': None,
            'end_coordinates': None,
            'raw_data': {}
        }

        # Обрабатываем координаты из SHR
        if messages.get('shr') and isinstance(messages['shr'], str):
            shr = messages['shr']

            # Ищем DEP координаты
            dep_match = re.search(r'DEP/\s*([0-9NSWECВСЗЮ]+)', shr)
            if dep_match:
                dep_coords = self.parse_coordinates_universal(dep_match.group(1))
                result['raw_data']['shr_dep'] = dep_coords
                result['start_coordinates'] = dep_coords

            # Ищем DEST координаты
            dest_match = re.search(r'DEST/\s*([0-9NSWECВСЗЮ]+)', shr)
            if dest_match:
                dest_coords = self.parse_coordinates_universal(dest_match.group(1))
                result['raw_data']['shr_dest'] = dest_coords
                result['end_coordinates'] = dest_coords

            # Ищем координаты в других форматах
            if not result['start_coordinates']:
                # Формат: DEP-координаты-...
                dep_alt_match = re.search(r'DEP-\s*([0-9NSWECВСЗЮ]+)', shr)
                if dep_alt_match:
                    dep_coords = self.parse_coordinates_universal(dep_alt_match.group(1))
                    result['raw_data']['shr_dep_alt'] = dep_coords
                    result['start_coordinates'] = dep_coords

        # Обрабатываем координаты из DEP
        if messages.get('dep') and isinstance(messages['dep'], str):
            dep = messages['dep']

            # Ищем ADEPZ координаты
            adepz_match = re.search(r'-ADEPZ\s+([0-9NSWECВСЗЮ]+)', dep)
            if adepz_match:
                adepz_coords = self.parse_coordinates_universal(adepz_match.group(1))
                result['raw_data']['dep_adepz'] = adepz_coords
                if not result['start_coordinates']:
                    result['start_coordinates'] = adepz_coords

            # Ищем координаты в других форматах DEP
            coord_matches = re.findall(r'[NS]\d+[EW]\d+', dep)
            for coord_match in coord_matches:
                coords = self.parse_coordinates_universal(coord_match)
                if coords and not result['start_coordinates']:
                    result['start_coordinates'] = coords
                    result['raw_data']['dep_auto'] = coords
                    break

        # Обрабатываем координаты из ARR
        if messages.get('arr') and isinstance(messages['arr'], str):
            arr = messages['arr']

            # Ищем ADARRZ координаты
            adarrz_match = re.search(r'-ADARRZ\s+([0-9NSWE]+)', arr)
            if adarrz_match:
                adarrz_coords = self.parse_coordinates_universal(adarrz_match.group(1))
                result['raw_data']['arr_adarrz'] = adarrz_coords
                if not result['end_coordinates']:
                    result['end_coordinates'] = adarrz_coords

            # Ищем координаты в других форматах ARR
            coord_matches = re.findall(r'[NS]\d+[EW]\d+', arr)
            for coord_match in coord_matches:
                coords = self.parse_coordinates_universal(coord_match)
                if coords and not result['end_coordinates']:
                    result['end_coordinates'] = coords
                    result['raw_data']['arr_auto'] = coords
                    break

        return result

    def parse_time_string(self, time_str):
        """Парсинг строки времени в объект time"""
        if not time_str or not isinstance(time_str, str):
            return None

        try:
            # Удаляем все нецифровые символы
            digits = re.sub(r'\D', '', time_str)

            if len(digits) == 4:
                # Формат HHMM
                hours = int(digits[:2])
                minutes = int(digits[2:4])
                if 0 <= hours < 24 and 0 <= minutes < 60:
                    return time(hours, minutes)
            elif len(digits) == 3:
                # Формат HMM
                hours = int(digits[0])
                minutes = int(digits[1:3])
                if 0 <= hours < 24 and 0 <= minutes < 60:
                    return time(hours, minutes)
            elif len(digits) == 2:
                # Формат MM
                minutes = int(digits)
                if 0 <= minutes < 60:
                    return time(0, minutes)
            elif len(digits) == 1:
                # Формат M
                minutes = int(digits)
                if 0 <= minutes < 60:
                    return time(0, minutes)
        except (ValueError, TypeError):
            pass

        return None

    def parse_date_string(self, date_str):
        """Парсинг строки дата в объект date"""
        if not date_str or not isinstance(date_str, str):
            return None

        try:
            # Удаляем все нецифровые символы
            digits = re.sub(r'\D', '', date_str)

            if len(digits) == 6:
                # Формат DDMMYY
                day = int(digits[:2])
                month = int(digits[2:4])
                year = 2000 + int(digits[4:6])
                if 1 <= day <= 31 and 1 <= month <= 12:
                    return date(year, month, day)
            elif len(digits) == 4:
                # Формат DDMM - используем текущий год
                day = int(digits[:2])
                month = int(digits[2:4])
                year = datetime.now().year
                if 1 <= day <= 31 and 1 <= month <= 12:
                    return date(year, month, day)
        except (ValueError, TypeError):
            pass

        return None

    def extract_flight_times(self, messages: Dict[str, str]) -> Dict[str, Any]:
        """Извлечение времени полета из всех источников"""
        result = {
            'shr': {},
            'dep': {},
            'arr': {},
            'final': {}
        }

        # Обрабатываем SHR
        if messages.get('shr') and isinstance(messages['shr'], str):
            shr = messages['shr']

            # Ищем временные метки ZZZZ
            if re.search(r'-ZZZZ(\d{4})', shr):
                matches = re.findall(r'-ZZZZ(\d{4})', shr)
                result['shr']['zzzz_times'] = matches

                if len(matches) >= 2:
                    result['shr']['end_time'] = self.parse_time_string(matches[1])
                    result['shr']['start_time'] = self.parse_time_string(matches[0])
                elif len(matches) == 1:
                    result['shr']['start_time'] = self.parse_time_string(matches[0])

            # Ищем дату полета DOF
            if re.search(r'DOF/(\d{6})', shr):
                match = re.search(r'DOF/(\d{6})', shr)
                date_str = match.group(1)
                year = int(date_str[0:2]) + 2000
                month = int(date_str[2:4])
                day = int(date_str[4:6])
                try:
                    flight_date = datetime(year, month, day)
                    result['shr']['flight_date'] = flight_date.strftime('%Y-%m-%d')
                except ValueError:
                    result['shr']['flight_date'] = None

        # Обрабатываем DEP
        if messages.get('dep') and isinstance(messages['dep'], str):
            dep = messages['dep']

            # Время вылета
            if re.search(r'-ATD\s+(\d{4})', dep):
                match = re.search(r'-ATD\s+(\d{4})', dep)
                result['dep']['atd_time'] = self.parse_time_string(match.group(1))

            # Дата вылета
            if re.search(r'-ADD\s+(\d{6})', dep):
                match = re.search(r'-ADD\s+(\d{6})', dep)
                result['dep']['add_date'] = self.parse_date_string(match.group(1))

        # Обрабатываем ARR
        if messages.get('arr') and isinstance(messages['arr'], str):
            arr = messages['arr']

            # Время прибытия
            if re.search(r'-ATA\s+(\d{4})', arr):
                match = re.search(r'-ATA\s+(\d{4})', arr)
                result['arr']['ata_time'] = self.parse_time_string(match.group(1))

            # Дата прибытия
            if re.search(r'-ADA\s+(\d{6})', arr):
                match = re.search(r'-ADA\s+(\d{6})', arr)
                result['arr']['ada_date'] = self.parse_date_string(match.group(1))

        # Формируем финальный результат
        result['final'] = {
            'start_time_shr': result['shr'].get('start_time'),
            'end_time_shr': result['shr'].get('end_time'),
            'date_shr': result['shr'].get('flight_date'),
            'start_time_dep': result['dep'].get('atd_time'),
            'date_dep': result['dep'].get('add_date'),
            'end_time_arr': result['arr'].get('ata_time'),
            'date_arr': result['arr'].get('ada_date'),
        }

        # Заполняем отсутствующие значения
        final = result['final']

        # Для дат
        if not final['date_dep'] and final['date_arr']:
            final['date_dep'] = final['date_arr']
        elif not final['date_arr'] and final['date_dep']:
            final['date_arr'] = final['date_dep']
        elif not final['date_dep'] and not final['date_arr'] and final['date_shr']:
            final['date_dep'] = final['date_shr']
            final['date_arr'] = final['date_shr']

        # Для времени
        if not final['start_time_dep'] and final['end_time_arr']:
            final['start_time_dep'] = final['end_time_arr']
        elif not final['end_time_arr'] and final['start_time_dep']:
            final['end_time_arr'] = final['start_time_dep']
        elif not final['start_time_dep'] and final['start_time_shr']:
            final['start_time_dep'] = final['start_time_shr']
        elif not final['end_time_arr'] and final['end_time_shr']:
            final['end_time_arr'] = final['end_time_shr']

        return result

    def extract_operator_info(self, text: str) -> Dict[str, Any]:
        """
        Извлечение информации об операторе с улучшенной обработкой.
        Сначала извлекает телефон из исходного текста, затем оператора (без телефона).
        Телефон должен соответствовать формату российского мобильного номера.
        """
        if not text or not isinstance(text, str):
            return {'operator': None, 'phone_number': None}

        result = {'operator': None, 'phone_number': None}

        # Сначала ищем все потенциальные номера телефонов в исходном тексте
        # Паттерн для поиска всех российских номеров
        all_phones_pattern = r'(?:\+7|8)[\s\-\(\)]*([\d\s\-\(\)]{10,})'
        all_matches = re.finditer(all_phones_pattern, text)
        
        # Собираем все найденные номера
        phone_candidates = []
        for match in all_matches:
            phone_text = match.group(0)
            # Очищаем от всех нецифровых символов
            cleaned = re.sub(r'[^\d]', '', phone_text)
            
            # Нормализуем: убираем начальную 8 или 7, оставляем 10 цифр
            if cleaned.startswith('8') and len(cleaned) == 11:
                cleaned = cleaned[1:]  # Убираем 8, оставляем 10 цифр
            elif cleaned.startswith('7') and len(cleaned) == 11:
                cleaned = cleaned[1:]  # Убираем 7, оставляем 10 цифр
            
            phone_candidates.append(cleaned)
        
        # Проверяем каждый номер на соответствие формату мобильного телефона
        for phone in phone_candidates:
            if self.is_valid_mobile_phone(phone):
                # Возвращаем в формате 8XXXXXXXXXX
                result['phone_number'] = '8' + phone
                break

        # Теперь извлекаем оператора (метод extract_operator уже удаляет телефоны)
        operator = self.extract_operator(text)
        if operator:
            result['operator'] = operator

        return result

    def is_valid_mobile_phone(self, phone: str) -> bool:
        """
        Проверка, является ли номер валидным российским мобильным телефоном.
        
        Args:
            phone: Номер телефона из 10 цифр (без кода страны)
        
        Returns:
            True, если номер соответствует формату мобильного телефона
        """
        if not phone or len(phone) != 10:
            return False
        
        # Проверяем, что номер состоит только из цифр
        if not phone.isdigit():
            return False
        
        # Проверяем, что код оператора (первые 3 цифры) в диапазоне 900-999
        # Российские мобильные операторы используют коды 900-999
        operator_code = phone[:3]
        try:
            code_num = int(operator_code)
            return 900 <= code_num <= 999
        except ValueError:
            return False

    def extract_flight_number(self, text: str) -> Optional[str]:
        """Извлечение номера рейса"""
        if not text or not isinstance(text, str):
            return None

        patterns = [
            r'FLIGHT\/([^\s]+)',
            r'FLC\/([^\s]+)',
            r'CALLSIGN\/([^\s]+)',
            r'(\b[A-Z]{2}\d{3,4}\b)',  # Пример: SU1234
            r'(\b[A-Z]{1,3}\d{2,5}\b)',  # Пример: A123, ABC1234
            r'(\b\d{3,6}\b)',  # Пример: 123, 12345
            r'RMK/.*?FLIGHT[:\s]*([^\n]+)',
            r'RMK/.*?CALLSIGN[:\s]*([^\n]+)',
            # Более общие паттерны
            r'(\b[A-Z]{1,4}\d{1,6}\b)',  # Любые буквы + цифры
            r'(\b\d{1,6}[A-Z]{1,4}\b)',  # Цифры + буквы
            r'(\b[A-Z]{2,4}\b)',  # Только буквы (коды авиакомпаний)
            r'(\b\d{2,6}\b)',  # Только цифры
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                flight_num = match.group(1).strip()
                # Проверяем, что это не слишком короткое число
                if len(flight_num) >= 2:
                    return flight_num

        # Если стандартные паттерны не сработали, попробуем более агрессивный поиск
        # Ищем любые комбинации букв и цифр
        aggressive_patterns = [
            r'([A-Z]{1,4}\d{1,6})',  # Буквы + цифры
            r'(\d{1,6}[A-Z]{1,4})',  # Цифры + буквы
            r'([A-Z]{2,6})',  # Только буквы
            r'(\d{2,6})',  # Только цифры
        ]

        for pattern in aggressive_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2 and len(match) <= 8:  # Разумная длина
                    return match.strip()

        return None

    def extract_altitude_info(self, text: str) -> Optional[str]:
        """
        Извлечение информации о высоте в исходном текстовом формате.
        
        Поддерживаемые форматы:
        - S0120/S0540 - диапазон высот по давлению 760 мм рт.ст.
        - M0021/M0043 - диапазон максимальной абсолютной высоты
        - M0065 - одиночное значение максимальной абсолютной высоты
        - K0200M0065 - скорость и высота
        
        Возвращает исходный текст высоты (например, "M0021/M0043")
        """
        if not text or not isinstance(text, str):
            return None

        # Паттерны для стандартного формата высоты
        standard_patterns = [
            r'\b([MS]\d{4}(?:/[MS]\d{4})?)\b',  # M0021/M0043, M0065, S0120/S0540
            r'K\d{4}([MS]\d{4})',  # K0200M0065
        ]

        for pattern in standard_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).upper()

        # Если не нашли стандартный формат, попробуем старые паттерны для обратной совместимости
        legacy_patterns = [
            r'ALT/([^\s]+)',
            r'ALTITUDE/([^\s]+)',
            r'LEVEL/([^\s]+)',
            r'FL/([^\s]+)',
            r'(\bFL\d{2,3}\b)',  # FL100, FL200
            r'(\b\d{3,4}\s*м\b)',  # 1000м, 2000м
            r'(\b\d{3,4}\s*M\b)',  # 1000M, 2000M
            r'(\b\d{3,4}\s*FT\b)',  # 1000FT, 2000FT
            r'RMK/.*?ALT[:\s]*([^\n]+)',
            r'RMK/.*?ALTITUDE[:\s]*([^\n]+)',
            r'RMK/.*?LEVEL[:\s]*([^\n]+)',
            # Более общие паттерны для высоты
            r'(\b\d{2,5}\s*м\b)',  # 100м, 5000м
            r'(\b\d{2,5}\s*M\b)',  # 100M, 5000M
            r'(\b\d{2,5}\s*FT\b)',  # 100FT, 5000FT
            r'(\b\d{2,5}\s*фут\b)',  # 100 фут, 5000 фут
            r'(\b\d{2,5}\s*feet\b)',  # 100 feet, 5000 feet
            r'(\b\d{2,5}\s*метров\b)',  # 100 метров, 5000 метров
            r'(\b\d{2,5}\s*метра\b)',  # 100 метра, 5000 метра
        ]

        for pattern in legacy_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                altitude_str = match.group(1).strip()
                return self.convert_altitude_to_meters(altitude_str)

        # Если стандартные паттерны не сработали, попробуем более агрессивный поиск
        # Ищем любые числа, которые могут быть высотой
        aggressive_patterns = [
            r'(\d{2,5})',  # Любые числа от 10 до 99999
        ]

        for pattern in aggressive_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                if match.isdigit():
                    num = int(match)
                    # Проверяем, что это может быть высота (от 10 до 50000 метров/футов)
                    if 10 <= num <= 50000:
                        return self.convert_altitude_to_meters(match)

        return None
    
    def extract_altitude_max(self, altitude_info: str) -> Optional[int]:
        """
        Извлечение максимального числового значения высоты из altitude_info.
        
        Обрабатывает форматы:
        - M0021/M0043 -> 43 (берем максимум из диапазона)
        - M0000/M0040 -> 40
        - M0065 -> 65
        - S0120/S0540 -> 540
        
        Возвращает числовое значение в десятках метров (без умножения на 10).
        Например, M0043 -> 43 (что означает 430 метров, но храним как 43).
        """
        if not altitude_info or not isinstance(altitude_info, str):
            return None
        
        try:
            # Паттерн для извлечения чисел из формата M0021/M0043 или S0120/S0540
            pattern = r'[MS](\d{4})'
            matches = re.findall(pattern, altitude_info, re.IGNORECASE)
            
            if matches:
                # Преобразуем все найденные значения в числа (убираем ведущие нули)
                numbers = [int(match) for match in matches]
                # Возвращаем максимальное значение
                return max(numbers)*10 # умножаем на 10 для перевода в метры
            
            return None
            
        except Exception as e:
            logging.debug(f"Ошибка извлечения altitude_max из '{altitude_info}': {e}")
            return None

    def convert_altitude_to_meters(self, altitude_str: str) -> Optional[str]:
        """
        Преобразует строку с высотой в числовое значение в метрах.
        Возвращает строку с числом (например "1500") или None если не удалось преобразовать.
        """
        if not altitude_str:
            return None

        try:
            # Очищаем строку от пробелов и лишних символов
            clean_str = str(altitude_str).strip().upper()

            # Удаляем все нецифровые символы кроме точки, запятой и минуса
            clean_str = re.sub(r'[^\d.,\-]', '', clean_str)

            # Заменяем запятую на точку для дробных чисел
            clean_str = clean_str.replace(',', '.')

            # Если строка пустая после очистки
            if not clean_str:
                return None

            # Пытаемся преобразовать в число
            altitude_value = float(clean_str)

            # Проверяем диапазон (от 0 до 20000 метров - разумный диапазон для БПЛА)
            if 0 <= altitude_value <= 20000:
                # Округляем до целого и возвращаем как строку
                return str(int(round(altitude_value)))
            else:
                return None

        except (ValueError, TypeError):
            return None


    def extract_remarks(self, text: str) -> Optional[str]:
        """Извлечение примечаний"""
        if not text or not isinstance(text, str):
            return None

        match = re.search(r'RMK\/([^\n]+)', text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

        return None

    def calculate_duration(self, start_time, end_time):
        """Расчет длительности полета в минутах"""
        if not start_time or not end_time:
            return None
        try:
            # Если start_time и end_time - объекты time
            if isinstance(start_time, time) and isinstance(end_time, time):
                start_dt = datetime.combine(date.today(), start_time)
                end_dt = datetime.combine(date.today(), end_time)
                # Если время окончания меньше времени начала, предполагаем, что полет закончился на следующий день
                if end_dt < start_dt:
                    end_dt = end_dt.replace(day=end_dt.day + 1)
                duration = end_dt - start_dt
                return int(duration.total_seconds() / 60)
        except Exception as e:
            logging.error(f"Ошибка расчета длительности: {e}")
        return None

    def process_flights(self, data, limit: int = None, batch_size: int = 1000):
        """Основной метод обработки данных о полетах"""
        if not self.connection:
            logging.error("Не инициализировано соединение с БД")
            return

        try:
            total_count = len(data)
            if limit:
                total_count = min(limit, total_count)
                logging.info(f"Обработка ограничена {limit} записями")

            self.total_records_in_file = total_count
            logging.info(f"Всего записей для обработки: {total_count}")
            self.update_progress(0, total_count, f"Всего записей для обработки: {total_count}")

            # Создание таблицы для результатов
            self.create_results_table()

            processed = 0
            start_time = time_module.time()
            batch_values = []

            for i, row in enumerate(data):
                if limit and processed >= limit:
                    break

                try:
                    if i % 5000 == 0:
                        self.update_progress(i, total_count, "Обработка записей")

                    # Обрабатываем запись и получаем значения для вставки
                    values = self.process_single_flight(row)
                    if values:
                        # Добавляем имя файла к значениям (36-й элемент) и time_of_day
                        time_of_day = self.calculate_time_of_day(values)
                        values_with_filename = values + (self.current_filename, 0, 'download', time_of_day)
                        batch_values.append(values_with_filename)
                        processed += 1

                    # Если накопился батч, вставляем
                    if len(batch_values) >= batch_size:
                        self.save_batch(batch_values)
                        batch_values = []
                        logging.info(f"Вставлено {processed} записей")

                except Exception as e:
                    logging.error(f"Ошибка обработки записи {i}: {e}")
                    continue

            # Вставляем оставшиеся записи
            if batch_values:
                self.save_batch(batch_values)

            self.update_progress(total_count, total_count, "Завершение обработки")
            total_time = time_module.time() - start_time
            logging.info(
                f"Обработка завершена. Время: {total_time:.1f} сек, скорость: {processed / total_time:.1f} записей/сек")

            # Логируем статистику заполненности полей
            self.log_field_statistics()

        except Exception as e:
            error_msg = f"Ошибка при обработке полетов: {e}"
            logging.error(error_msg)
            if self.connection:
                self.connection.rollback()

    def save_batch(self, batch_values):
        """
        Массовая вставка батча данных.
        
        Примечание: Используется parameterized query (с %s плейсхолдерами),
        что автоматически обеспечивает экранирование кавычек, апострофов и других 
        специальных символов для безопасной вставки в базу данных.
        """
        if not batch_values:
            return
        try:
            cursor = self.connection.cursor()
            
            # Индексы столбцов в batch_values для формирования UNIQ_STR
            column_indices = {
                'source_id': 0, 'sheet_name': 1, 'flight_id': 2, 'flight_number': 3,
                'aircraft_type_code': 4, 'aircraft_type_desc': 5,
                'region_id': 6, 'departure_region': 7, 'arrival_region': 8,
                'duration_min': 9, 'operator': 10, 'registration': 11, 'registration_all': 12,
                'departure_coords': 13, 'arrival_coords': 14,
                'departure_lat': 15, 'departure_lon': 16, 'arrival_lat': 17, 'arrival_lon': 18,
                'dof_date': 19, 'arrival_actual_date': 20, 'departure_actual_date': 21, 'planned_date': 22,
                'atd_time': 23, 'departure_actual_time': 24, 'ata_time': 25, 'arrival_actual_time': 26,
                'eet_info': 27, 'shr_message': 28, 'dep_message': 29, 'arr_message': 30, 
                'center_es_orvd': 31, 'altitude_info': 32, 'altitude_max': 33,
                'phone_number': 34, 'quantity': 35, 'sts': 36, 'filename': 37, 
                'hexagon_id': 38, 'prediction': 39, 'time_of_day': 40
            }
            
            # Добавляем UNIQ_STR к каждой записи
            batch_values_with_uniq = []
            for values in batch_values:
                # Формируем UNIQ_STR из указанных столбцов
                uniq_parts = []
                for col_name in self.UNIQ_STR_COLUMNS:
                    if col_name in column_indices:
                        idx = column_indices[col_name]
                        value = values[idx] if idx < len(values) else None
                        # Преобразуем значение в строку, обрабатывая None
                        uniq_parts.append(str(value) if value is not None else '')
                
                uniq_str = '|'.join(uniq_parts)
                # Добавляем uniq_str в конец кортежа
                batch_values_with_uniq.append(values + (uniq_str,))
            
            query = """
            INSERT INTO processed_flights (
    source_id, sheet_name, flight_id, flight_number, 
    aircraft_type_code, aircraft_type_desc,
    region_id, departure_region, arrival_region,
    duration_min, operator, registration, registration_all,
    departure_coords, arrival_coords,
    departure_lat, departure_lon, arrival_lat, arrival_lon,
    dof_date, arrival_actual_date, departure_actual_date, planned_date,
    atd_time, departure_actual_time, ata_time, arrival_actual_time,
    eet_info, shr_message, dep_message, arr_message, center_es_orvd, altitude_info, altitude_max,
    phone_number, quantity, sts, filename, hexagon_id, prediction, time_of_day, uniq_str
) VALUES (
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s
)
            """
            cursor.executemany(query, batch_values_with_uniq)
            self.connection.commit()
            logging.info(f"Успешно вставлено {cursor.rowcount} записей.")
        except Exception as e:
            self.connection.rollback()
            logging.error(f"Ошибка при batch-вставке: {e}")
            # Для диагностики
            if batch_values:
                logging.error(f"Количество значений в кортеже: {len(batch_values[0])}, ожидается 41")
                logging.error(f"Пример значений: {batch_values[0]}")
            raise
        finally:
            cursor.close()

    def remove_duplicates_by_uniq_str(self):
        """
        Удаление дубликатов из таблицы processed_flights по полю uniq_str.
        
        Оставляет только запись с минимальным id среди дубликатов.
        Выводит статистику удаленных записей.
        """
        try:
            cursor = self.connection.cursor(dictionary=True)
            
            # Сначала получаем статистику дубликатов
            cursor.execute("""
                SELECT uniq_str, COUNT(*) as count
                FROM processed_flights
                WHERE uniq_str IS NOT NULL AND uniq_str != ''
                GROUP BY uniq_str
                HAVING count > 1
            """)
            duplicates = cursor.fetchall()
            
            if not duplicates:
                logging.info("Дубликаты не найдены")
                return 0
            
            total_duplicate_groups = len(duplicates)
            total_duplicate_records = sum(d['count'] for d in duplicates)
            logging.info(f"Найдено {total_duplicate_groups} групп дубликатов, всего {total_duplicate_records} записей")
            
            # Удаляем дубликаты, оставляя запись с минимальным id
            delete_query = """
                DELETE t1 FROM processed_flights t1
                INNER JOIN (
                    SELECT uniq_str, MIN(id) as min_id
                    FROM processed_flights
                    WHERE uniq_str IS NOT NULL AND uniq_str != ''
                    GROUP BY uniq_str
                    HAVING COUNT(*) > 1
                ) t2 ON t1.uniq_str = t2.uniq_str
                WHERE t1.id > t2.min_id
            """
            cursor.execute(delete_query)
            deleted_count = cursor.rowcount
            self.connection.commit()
            
            logging.info(f"Удалено {deleted_count} дубликатов из таблицы processed_flights")
            logging.info(f"Оставлено {total_duplicate_groups} уникальных записей (по одной из каждой группы дубликатов)")
            
            return deleted_count
            
        except Exception as e:
            self.connection.rollback()
            logging.error(f"Ошибка при удалении дубликатов: {e}")
            raise
        finally:
            if cursor:
                cursor.close()

    def create_results_table(self):
        """Создание таблицы для результатов обработки"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SHOW TABLES LIKE 'processed_flights'")
            if cursor.fetchone():
                # Проверяем наличие поля prediction
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'prediction'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN prediction ENUM('download', 'prediction') NOT NULL DEFAULT 'download'")
                    self.connection.commit()
                    logging.info("Добавлено поле prediction в таблицу processed_flights")

                # Проверяем наличие новых полей для исходных сообщений
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'shr_message'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN shr_message TEXT COMMENT 'Исходное сообщение SHR'")
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN dep_message TEXT COMMENT 'Исходное сообщение DEP'")
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN arr_message TEXT COMMENT 'Исходное сообщение ARR'")
                    self.connection.commit()
                    logging.info("Добавлены поля shr_message, dep_message, arr_message в таблицу processed_flights")

                # Проверяем наличие поля remarks и удаляем его
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'remarks'")
                if cursor.fetchone():
                    cursor.execute("ALTER TABLE processed_flights DROP COLUMN remarks")
                    self.connection.commit()
                    logging.info("Удалено поле remarks из таблицы processed_flights")

                # Проверяем наличие поля registration_all
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'registration_all'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN registration_all TEXT COMMENT 'Все регистрационные номера БПЛА через запятую'")
                    self.connection.commit()
                    logging.info("Добавлено поле registration_all в таблицу processed_flights")

                # Проверяем наличие поля time_of_day
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'time_of_day'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN time_of_day ENUM('утро','день','вечер','ночь') DEFAULT 'день' COMMENT 'Время суток с учетом UTC региона'")
                    self.connection.commit()
                    logging.info("Добавлено поле time_of_day в таблицу processed_flights")

                # Проверяем наличие поля sts
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'sts'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN sts VARCHAR(50) DEFAULT NULL COMMENT 'Причина особого отношения (FFR, SAR, STATE и др.)'")
                    self.connection.commit()
                    logging.info("Добавлено поле sts в таблицу processed_flights")

                # Проверяем наличие поля uniq_str
                cursor.execute("SHOW COLUMNS FROM processed_flights LIKE 'uniq_str'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE processed_flights ADD COLUMN uniq_str VARCHAR(500) DEFAULT NULL COMMENT 'Уникальная строка для анализа дубликатов'")
                    cursor.execute("ALTER TABLE processed_flights ADD INDEX idx_uniq_str (uniq_str)")
                    self.connection.commit()
                    logging.info("Добавлено поле uniq_str в таблицу processed_flights")

                logging.info("Таблица processed_flights уже существует и обновлена")
                
             
                return

            cursor.execute("""

                CREATE TABLE `processed_flights` (
                  `id` int(11) NOT NULL AUTO_INCREMENT,
                  `source_id` varchar(255) DEFAULT NULL COMMENT 'ID из исходных данных',
                  `sheet_name` varchar(100) DEFAULT NULL COMMENT 'Название листа в Excel',
                  `flight_id` varchar(255) DEFAULT NULL COMMENT 'Идентификатор полета SID',
                  `flight_number` varchar(100) DEFAULT NULL COMMENT 'Номер рейса',
                  `aircraft_type_code` varchar(10) DEFAULT NULL COMMENT 'Код типа ВС из TYP/ (BLA, AER, SHAR)',
                  `aircraft_type_desc` varchar(100) DEFAULT NULL COMMENT 'Расшифровка типа ВС',
                  `region_id` smallint(2) NOT NULL COMMENT 'Идентификатор региона',
                  `departure_region` varchar(100) DEFAULT NULL COMMENT 'Регион вылета',
                  `arrival_region` varchar(100) DEFAULT NULL COMMENT 'Регион прибытия',
                  `duration_min` int(11) DEFAULT NULL COMMENT 'Длительность полета в минутах',
                  `operator` varchar(100) DEFAULT NULL COMMENT 'Оператор (ФИО или организация)',
                  `registration` varchar(255) DEFAULT NULL COMMENT 'Регистрационный номер ВС',
                  `registration_all` text COMMENT 'Все регистрационные номера БПЛА через запятую',
                  `departure_coords` varchar(100) DEFAULT NULL COMMENT 'Координаты вылета в исходном формате',
                  `arrival_coords` varchar(100) DEFAULT NULL COMMENT 'Координаты прибытия в исходном формате',
                  `departure_lat` float DEFAULT NULL COMMENT 'Широта вылета в десятичных градусах',
                  `departure_lon` float DEFAULT NULL COMMENT 'Долгота вылета в десятичных градусах',
                  `arrival_lat` float DEFAULT NULL COMMENT 'Широта прибытия в десятичных градусах',
                  `arrival_lon` float DEFAULT NULL COMMENT 'Долгота прибытия в десятичных градусах',
                  `dof_date` date DEFAULT NULL COMMENT 'Дата полета (устаревшее поле, используйте planned_date, departure_actual_date, arrival_actual_date)',
                  `arrival_actual_date` date DEFAULT NULL COMMENT 'Фактическая дата прибытия из ADA',
                  `departure_actual_date` date DEFAULT NULL COMMENT 'Фактическая дата вылета из ADD',
                  `planned_date` date DEFAULT NULL COMMENT 'Дата полета из DOF/',
                  `atd_time` time DEFAULT NULL COMMENT 'Время вылета (устаревшее поле, используйте departure_actual_time)',
                  `departure_actual_time` time DEFAULT NULL COMMENT 'Фактическое время вылета из ATD',
                  `ata_time` time DEFAULT NULL COMMENT 'Время прибытия (устаревшее поле, используйте arrival_actual_time)',
                  `arrival_actual_time` time DEFAULT NULL COMMENT 'Фактическое время прибытия из ATA',
                  `eet_info` varchar(100) DEFAULT NULL COMMENT 'Расчетное время полета',
                  `center_es_orvd` varchar(100) DEFAULT NULL COMMENT 'Центр ЕС ОрВД',
                  `altitude_info` varchar(100) DEFAULT NULL COMMENT 'Информация о высоте (текстовый формат, например M0021/M0043)',
                  `altitude_max` int(11) DEFAULT NULL COMMENT 'Максимальная высота в десятках метров (числовое значение, например 43)',
                  `phone_number` varchar(255) DEFAULT NULL COMMENT 'Номер телефона оператора',
                  `quantity` int(11) DEFAULT NULL COMMENT 'Количество БПЛА',
                  `sts` varchar(50) DEFAULT NULL COMMENT 'Причина особого отношения (FFR, SAR, STATE и др.)',
                  `processed_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  `filename` varchar(100) DEFAULT NULL COMMENT 'Имя исходного файла',
                  `hexagon_id` int(11) NOT NULL DEFAULT '0',
                  `prediction` enum('download','prediction') NOT NULL DEFAULT 'download',
                  `shr_message` text COMMENT 'Исходное сообщение SHR',
                  `dep_message` text COMMENT 'Исходное сообщение DEP',
                  `arr_message` text COMMENT 'Исходное сообщение ARR',
                  `time_of_day` enum('утро','день','вечер','ночь') DEFAULT 'день' COMMENT 'Время суток с учетом UTC региона',
                  `uniq_str` varchar(500) DEFAULT NULL COMMENT 'Уникальная строка для анализа дубликатов',
                  PRIMARY KEY (`id`),
                  KEY `idx_departure_region` (`departure_region`),
                  KEY `idx_arrival_region` (`arrival_region`),
                  KEY `idx_operator` (`operator`(20)),
                  KEY `idx_dof_date` (`dof_date`),
                  KEY `idx_sheet_name` (`sheet_name`),
                  KEY `idx_filename` (`filename`),
                  KEY `idx_departure_coords` (`departure_lat`,`departure_lon`),
                  KEY `dof_date` (`dof_date`,`prediction`),
                  KEY `idx_processed_flights_coords` (`departure_lat`,`departure_lon`),
                  KEY `idx_uniq_str` (`uniq_str`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='Таблица обработанных данных о полетах БПЛА' 
            """)
            self.connection.commit()
            logging.info("Таблица результатов создана")
            
          
        except Exception as e:
            logging.error(f"Ошибка создания таблицы результатов: {e}")
        finally:
            if cursor:
                cursor.close()

   

    def log_field_statistics(self): 
        """Логирование статистики заполненности полей"""
        try:
            cursor = self.connection.cursor(dictionary=True)

            fields_to_check = [
                'operator',
                # 'registration',
                # 'registration_all',
                # 'quantity',
                # 'flight_id',
                # 'flight_number',
                # 'aircraft_type_code',
                # 'aircraft_type_desc',
                # 'region_id',
                # 'departure_region',
                # 'arrival_region',
                # 'duration_min',
                # 'departure_coords',
                # 'arrival_coords',
                # 'departure_lat',
                # 'departure_lon',
                # 'arrival_lat',
                # 'arrival_lon',
                # 'dof_date',
                # 'arrival_actual_date',
                # 'departure_actual_date',
                # 'planned_date',
                # 'atd_time',
                # 'departure_actual_time',
                # 'ata_time',
                # 'arrival_actual_time',
                # 'eet_info',
                # 'shr_message',
                # 'dep_message',
                # 'arr_message',
                # 'center_es_orvd',
                # 'altitude_info',
                # 'altitude_max',
                # 'phone_number',
                # 'sts',
                # 'processed_at',
                # 'filename',
                # 'hexagon_id',
                # 'prediction'
            ]

            stats = {}
            cursor.execute("SELECT COUNT(*) as count FROM processed_flights")
            total_records = cursor.fetchone()['count']

            for field in fields_to_check:
                cursor.execute(f"SELECT COUNT(*) as count FROM processed_flights WHERE {field} IS NOT NULL")
                non_null_count = cursor.fetchone()['count']
                stats[field] = {
                    'non_null': non_null_count,
                    'null': total_records - non_null_count,
                    'percentage': (non_null_count / total_records * 100) if total_records > 0 else 0
                }

            logging.info("=== СТАТИСТИКА ЗАПОЛНЕННОСТИ ПОЛЕЙ ===")
            self.update_progress(99, 100, f"СТАТИСТИКА ЗАПОЛНЕННОСТИ ПОЛЕЙ ")
            for field, data in stats.items():
                logging.info(f"{field}: {data['non_null']}/{total_records} ({data['percentage']:.1f}%) заполнено")
                self.update_progress(99, 100,
                                     f"{field}: {data['non_null']}/{total_records} ({data['percentage']:.1f}%) заполнено")

        except Exception as e:
            logging.error(f"Ошибка при получении статистики: {e}")
        finally:
            if cursor:
                cursor.close()

    def close_connection(self):
        """Закрытие соединения с БД"""
        if self.connection:
            self.connection.close()
            logging.info("Соединение с базой данных закрыто")

    def process_all_files(self, force_reprocess=False):
        """Совместимость: просто вызывает обработку непроцессированных файлов.
        
        Args:
            force_reprocess (bool): Если True, переобрабатывает все файлы, даже уже обработанные
        """
        self.process_unprocessed_files(force_reprocess)
    


if __name__ == "__main__":
    """Функция для запуска модуля отдельно"""
    processor = FlightDataProcessor()
    processor.process_all_files()
    # processor.remove_duplicates_by_uniq_str()
    